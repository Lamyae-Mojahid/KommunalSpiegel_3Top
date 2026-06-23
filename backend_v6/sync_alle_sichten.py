#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KommunalSpiegel · Sync-Agent v6 · Alle Sichten · Verifizierte Endpunkte

VERIFIZIERTE DATENQUELLEN (Stand Juni 2026):
───────────────────────────────────────────────────────────────────────
S4  BNetzA Ladesäulenregister XLSX
    URL:  https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/
          Sachgebiete/Energie/Unternehmen_Institutionen/E_Mobilitaet/
          Ladesaeulenregister.xlsx?__blob=publicationFile
    Auch: https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/
          Sachgebiete/Energie/Unternehmen_Institutionen/E_Mobilitaet/
          Ladesaeulenregister.csv?__blob=publicationFile
    Status: Öffentlich, CC BY 4.0, kein Key, täglich aktualisiert
    Hinweis: Direktdownload funktioniert im Browser. Wenn Python-Download
             gesperrt, manuell herunterladen und --file übergeben.

S5  BNetzA Breitbandatlas XLSX (Gemeindeebene)
    URL:  https://data.bundesnetzagentur.de/Bundesnetzagentur/GIGA/DE/
          Breitbandatlas/Downloads/bba_12_2025.xlsx
    Auch: https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html
          → Filter: Festnetz, Tabelle, aktuell
    Status: Öffentlich, CC BY 4.0, kein Key, halbjährlich (Juni + Dezember)
    Inhalt: Spalte "Gemeinde", "Anteil_GBit_FTTB_H_%", Download-Klassen

S6  BNetzA Mobilfunk-Monitoring XLSX (Gemeindeebene)
    URL:  https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/
          Auswertung_Mobilfunkmonitoring.xlsx?__blob=publicationFile&v=11
    Auch: https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html
          → Filter: Mobilfunk, Tabelle, aktuell
    Status: Öffentlich, CC BY 4.0, kein Key, halbjährlich
    Inhalt: 4G/5G-Flächenversorgung in % nach Gemeinde + Netzbetreiber

S3  Google Street View Metadata API
    URL:  https://maps.googleapis.com/maps/api/streetview/metadata
    Key:  GOOGLE_MAPS_API_KEY (kostenlos, Metadata-Anfragen = 0 USD)
    Erstellen: https://console.cloud.google.com →
               APIs & Services → Credentials → API Key erstellen →
               "Street View Static API" aktivieren
    Hinweis: Billing-Konto nötig, aber Metadata kostet nichts

S7  OZG-Umsetzungsstand (Sachsen-Anhalt Serviceportal + manuelle CSV)
    URL:  https://service.sachsen-anhalt.de  (öffentlich, kein Key)
    Alt:  https://www.digitale-verwaltung.de/Webs/DV/DE/
          (kein direkter CSV-API-Endpunkt, Scraping oder manueller Download)
    Status: Benchmark-CSV bleibt primäre Quelle bis offizielle API verfügbar

S8  Meta Graph API
    URL:  https://graph.facebook.com/v19.0/search
    Key:  META_ACCESS_TOKEN
    Erstellen: https://developers.facebook.com →
               My Apps → Create App → Graph API Explorer →
               User Access Token (60 Tage gültig)
    Hinweis: Für Produktion: App Review beantragen (1-4 Wochen)

S2/9/10  OSM Overpass API
    URL:  https://overpass-api.de/api/interpreter
    Key:  kein Key nötig
    Status: Öffentlich, kostenlos, aus GitHub Actions erreichbar

S1  Claude API (IGEK/ISEK PDF-Parsing)
    URL:  https://api.anthropic.com/v1/messages
    Key:  ANTHROPIC_API_KEY
    Erstellen: https://console.anthropic.com → API Keys
───────────────────────────────────────────────────────────────────────

Verwendung:
  pip install -r requirements.txt
  cp .env.example .env       # Keys eintragen
  python sync_alle_sichten.py --alle
  python sync_alle_sichten.py --sicht 4 --file ~/Downloads/Ladesaeulenregister.xlsx
  python sync_alle_sichten.py --sicht 5 --file ~/Downloads/bba_12_2025.xlsx
  python sync_alle_sichten.py --sicht 6 --file ~/Downloads/Auswertung_Mobilfunkmonitoring.xlsx
"""

from __future__ import annotations
import argparse, base64, csv, io, json, math, os, re, sys
import tempfile, time
from dataclasses import dataclass, asdict
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urljoin

import requests

# ─── ENV ──────────────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

GOOGLE_KEY    = os.getenv("GOOGLE_MAPS_API_KEY", "")
META_TOKEN    = os.getenv("META_ACCESS_TOKEN", "")
ANTHROPIC_KEY = os.getenv("ANTHROPIC_API_KEY", "")
SUPABASE_URL  = os.getenv("SUPABASE_URL", "") or os.getenv("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

# ─── VERZEICHNISSE ────────────────────────────────────────────────────────────
DATA_DIR  = Path("data")
CACHE_DIR = DATA_DIR / "cache"
RAW_DIR   = DATA_DIR / "raw"
OUT_DIR   = DATA_DIR / "output"
GEO_CACHE = CACHE_DIR / "gemeindegrenzen.json"

for d in (CACHE_DIR, RAW_DIR, OUT_DIR):
    d.mkdir(parents=True, exist_ok=True)

UA = "KommunalSpiegel/6.0 Hochschule-Merseburg studentisches-Projekt"
NOMINATIM_DELAY = 1.2
VERSION = "6.0.0"

# ─── VERIFIZIERTE DOWNLOAD-URLS ───────────────────────────────────────────────
# Quelle: https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/start.html
# Stand: Juni 2026 — URL-Muster bleibt stabil, Versionsnummer im Dateinamen kann wechseln
S4_URLS = [
    # XLSX (primär, ~26 MB, Stand April 2026)
    "https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/Sachgebiete/Energie/"
    "Unternehmen_Institutionen/E_Mobilitaet/Ladesaeulenregister.xlsx?__blob=publicationFile",
    # CSV-Variante (~46 MB)
    "https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/Sachgebiete/Energie/"
    "Unternehmen_Institutionen/E_Mobilitaet/Ladesaeulenregister.csv?__blob=publicationFile",
]

# Quelle: https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html
# Stand: Dezember 2025 — Dateiname enthält Datum, muss bei Aktualisierung angepasst werden
S5_URLS = [
    # Dateiname = bba_MM_YYYY.xlsx — aktueller Stand Ende 2025
    "https://data.bundesnetzagentur.de/Bundesnetzagentur/GIGA/DE/"
    "Breitbandatlas/Downloads/bba_12_2025.xlsx",
]
# Fallback-URL für S5 wenn Datei sich umbenennt:
S5_PAGE = "https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html"

# Quelle: https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/
# ?v=11 = Versionsnummer, kann sich ändern
S6_URLS = [
    "https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/"
    "Auswertung_Mobilfunkmonitoring.xlsx?__blob=publicationFile&v=11",
]
S6_PAGE = "https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html"


# ─── KOMMUNEN ─────────────────────────────────────────────────────────────────
KOMMUNEN: List[Dict[str, Any]] = [
    {"name":"Aken",                  "lk":"Anhalt-Bitterfeld", "ew":7200,  "lat":51.8533,"lng":12.0446},
    {"name":"Köthen",                "lk":"Anhalt-Bitterfeld", "ew":26500, "lat":51.7510,"lng":11.9729},
    {"name":"Muldestausee",          "lk":"Anhalt-Bitterfeld", "ew":7800,  "lat":51.6895,"lng":12.3165},
    {"name":"Osternienburger Land",  "lk":"Anhalt-Bitterfeld", "ew":8700,  "lat":51.7203,"lng":12.1472},
    {"name":"Raguhn-Jeßnitz",        "lk":"Anhalt-Bitterfeld", "ew":8200,  "lat":51.7131,"lng":12.2861},
    {"name":"Sandersdorf-Brehna",    "lk":"Anhalt-Bitterfeld", "ew":14200, "lat":51.6124,"lng":12.2280},
    {"name":"Südliches Anhalt",      "lk":"Anhalt-Bitterfeld", "ew":16100, "lat":51.6580,"lng":11.9055},
    {"name":"Zerbst/Anhalt",         "lk":"Anhalt-Bitterfeld", "ew":20700, "lat":51.9679,"lng":12.0898},
    {"name":"Zörbig",                "lk":"Anhalt-Bitterfeld", "ew":8600,  "lat":51.6249,"lng":12.1249},
    {"name":"Hohenmölsen",           "lk":"Burgenlandkreis",   "ew":9200,  "lat":51.1588,"lng":12.0944},
    {"name":"Lützen",                "lk":"Burgenlandkreis",   "ew":8100,  "lat":51.2566,"lng":12.1407},
    {"name":"Teuchern",              "lk":"Burgenlandkreis",   "ew":7100,  "lat":51.1508,"lng":12.0175},
    {"name":"Allstedt",              "lk":"Mansfeld-Südharz",  "ew":8300,  "lat":51.3979,"lng":11.4119},
    {"name":"Arnstein",              "lk":"Mansfeld-Südharz",  "ew":7900,  "lat":51.5588,"lng":11.3758},
    {"name":"Eisleben",              "lk":"Mansfeld-Südharz",  "ew":23800, "lat":51.5294,"lng":11.5492},
    {"name":"Gerbstedt",             "lk":"Mansfeld-Südharz",  "ew":7300,  "lat":51.6252,"lng":11.6141},
    {"name":"Hettstedt",             "lk":"Mansfeld-Südharz",  "ew":13400, "lat":51.6488,"lng":11.5022},
    {"name":"Mansfeld",              "lk":"Mansfeld-Südharz",  "ew":7700,  "lat":51.5968,"lng":11.4632},
    {"name":"Seegebiet Mansfelder Land","lk":"Mansfeld-Südharz","ew":8400, "lat":51.5699,"lng":11.7044},
    {"name":"Südharz",               "lk":"Mansfeld-Südharz",  "ew":8600,  "lat":51.5117,"lng":10.9705},
    {"name":"Bad Dürrenberg",        "lk":"Saalekreis",        "ew":9800,  "lat":51.2965,"lng":12.0645},
    {"name":"Bad Lauchstädt",        "lk":"Saalekreis",        "ew":9100,  "lat":51.3875,"lng":11.8714},
    {"name":"Braunsbedra",           "lk":"Saalekreis",        "ew":10600, "lat":51.2872,"lng":11.8843},
    {"name":"Kabelsketal",           "lk":"Saalekreis",        "ew":11200, "lat":51.4522,"lng":12.0089},
    {"name":"Landsberg",             "lk":"Saalekreis",        "ew":9700,  "lat":51.5219,"lng":11.9877},
    {"name":"Leuna",                 "lk":"Saalekreis",        "ew":14500, "lat":51.3286,"lng":12.0032,"ags":"15088075"},
    {"name":"Mücheln (Geiseltal)",   "lk":"Saalekreis",        "ew":8600,  "lat":51.2994,"lng":11.7996},
    {"name":"Petersberg",            "lk":"Saalekreis",        "ew":13100, "lat":51.5428,"lng":11.9875},
    {"name":"Querfurt",              "lk":"Saalekreis",        "ew":12800, "lat":51.3803,"lng":11.5897},
    {"name":"Salzatal",              "lk":"Saalekreis",        "ew":8800,  "lat":51.4610,"lng":11.7783},
    {"name":"Schkopau",              "lk":"Saalekreis",        "ew":9300,  "lat":51.3860,"lng":12.0049},
    {"name":"Teutschenthal",         "lk":"Saalekreis",        "ew":8900,  "lat":51.4589,"lng":11.7875},
    {"name":"Wettin-Löbejün",        "lk":"Saalekreis",        "ew":11800, "lat":51.5490,"lng":11.8855},
]

# AGS für S7 (OZG)
AGS = {
    "Leuna":"15088075","Bad Dürrenberg":"15088010","Bad Lauchstädt":"15088015",
    "Braunsbedra":"15088025","Kabelsketal":"15088050","Landsberg":"15088065",
    "Mücheln (Geiseltal)":"15088090","Petersberg":"15088095","Querfurt":"15088100",
    "Salzatal":"15088105","Schkopau":"15088110","Teutschenthal":"15088120",
    "Wettin-Löbejün":"15088125","Eisleben":"15087025","Hettstedt":"15087035",
    "Mansfeld":"15087050","Allstedt":"15087005","Gerbstedt":"15087030",
    "Seegebiet Mansfelder Land":"15087060",
}

# Benchmark-Daten (manuelle Erhebung 2024) — Fallback für alle Sichten
BENCHMARK: Dict[str, Dict[int, Any]] = {
    "Köthen":       {2:0.00,3:10.04,4:0.57,5:410.0, 6:{"o2":-12.39,"voda":-18.57,"tele":-14.13},7:{"rg":1.33,"ab2":33.33},8:0.62},
    "Muldestausee": {2:0.13,3:68.40,4:0.60,5:379.67,6:{"o2":-13.59,"voda":-10.98,"tele":-12.30},8:0.35},
    "Osternienburger Land":{2:0.13,3:66.67,5:18.28,6:{"o2":-10.23,"voda":-4.94,"tele":-6.71},7:{"rg":None,"ab2":38.0},8:1.52},
    "Sandersdorf-Brehna":{3:35.00,4:1.13,5:200.0,6:{"o2":-98,"voda":-90,"tele":-85},7:{"rg":1.70,"ab2":57.0},8:0.40},
    "Südliches Anhalt":{2:0.00,3:100.0,4:0.00,7:{"rg":1.37,"ab2":35.5}},
    "Hohenmölsen":  {4:0.30,7:{"rg":1.27,"ab2":27.77},8:0.07},
    "Allstedt":     {2:0.00,3:65.70,4:0.53,5:270.52,6:{"o2":-86.25,"voda":-91.29,"tele":-89.58},7:{"rg":0.98,"ab2":28.88},8:1.86},
    "Eisleben":     {2:0.22,3:75.0, 4:0.84,5:447.84,6:{"o2":-96.95,"voda":-94.03},7:{"rg":2.57,"ab2":60.71},8:0.71},
    "Gerbstedt":    {2:0.00,3:84.25,4:0.00,5:310.12,6:{"o2":-94.89,"voda":-100.78,"tele":-92.27},7:{"rg":1.04,"ab2":4.08},8:2.27},
    "Seegebiet Mansfelder Land":{2:0.00,3:78.70,4:0.94,5:559.51,6:{"o2":-87.43,"voda":-94.75,"tele":-88.16},7:{"rg":1.06,"ab2":10.43},8:2.48},
    "Bad Dürrenberg":{2:0.61,3:85.66,4:2.95,5:432.14,6:{"o2":-90.11,"voda":-98.29,"tele":-95.93},7:{"rg":0.571,"ab2":0.0},8:3.30},
    "Bad Lauchstädt":{2:0.11,3:87.30,4:0.89,5:550.80,6:{"o2":-88,"voda":-91.10,"tele":-87.40},7:{"rg":1.69,"ab2":60.87},8:0.89},
    "Braunsbedra":  {2:0.58,4:0.89,5:399.40,6:{"o2":-88.72,"voda":-96.48,"tele":-89.82}},
    "Kabelsketal":  {2:0.00,3:77.43,4:0.11,5:380.81,6:{"o2":-88.72,"voda":-96.48,"tele":-89.82},7:{"rg":1.31,"ab2":18.40},8:1.25},
    "Landsberg":    {2:0.00,3:81.10,4:1.02,5:400.06,6:{"o2":-79.74,"voda":-88.46,"tele":-80.19},7:{"rg":1.06,"ab2":6.12},8:0.96},
    "Leuna":        {2:0.41,3:93.01,4:3.59,5:447.67,6:{"o2":-90.40,"voda":-95.99,"tele":-81.86},7:{"rg":3.24,"ab2":84.85},8:5.66},
    "Mücheln (Geiseltal)":{2:0.32,3:72.0,4:0.21,5:510.38,6:{"o2":-95,"voda":-81.40,"tele":-90.60},7:{"rg":1.12,"ab2":23.08},8:0.61},
    "Querfurt":     {2:0.40,3:87.27,4:0.504,5:576.70,6:{"o2":-81.70,"voda":-69.60,"tele":-73.30},7:{"rg":1.50,"ab2":50.0},8:0.70},
    "Salzatal":     {2:0.02,3:83.55,4:0.88,5:443.20,6:{"o2":-92.30,"voda":-83.60,"tele":-86.30},7:{"rg":2.14,"ab2":66.67},8:0.05},
    "Wettin-Löbejün":{2:0.00,3:29.0,4:0.21,5:413.16,6:{"o2":-97.32,"voda":-93.07,"tele":-102.46},7:{"rg":1.15,"ab2":15.0},8:1.03},
}


# ─── LOGGING ──────────────────────────────────────────────────────────────────
def log(m): print(f"[{datetime.now().strftime('%H:%M:%S')}] {m}", flush=True)
def ok(m):  log(f"  ✓ {m}")
def warn(m):log(f"  ⚠ {m}")
def err(m): log(f"  ✗ {m}")
def fall(m):log(f"  ⑊ {m}  →  Benchmark-Fallback")


# ─── HILFSFUNKTIONEN ──────────────────────────────────────────────────────────
def norm_key(s):
    s = str(s or "").strip().lower()
    for a,b in {"ä":"ae","ö":"oe","ü":"ue","ß":"ss"}.items(): s=s.replace(a,b)
    return re.sub(r"[^a-z0-9]","",s)

def pf(v):
    """Parse deutschen Dezimalwert: '18,28 Mbit/s' → 18.28"""
    if v is None: return None
    if isinstance(v,(int,float)) and not isinstance(v,bool): return float(v)
    s = str(v).strip()
    # Einheit entfernen
    s = re.sub(r'[^0-9,.\-]','',s)
    if not s: return None
    if ',' in s and '.' in s:
        s = s.replace('.','').replace(',','.') if s.index('.')<s.index(',') else s.replace(',','')
    else:
        s = s.replace(',','.')
    try: return float(s)
    except: return None

def pick(row, cols):
    m = {norm_key(k):v for k,v in row.items()}
    for c in cols:
        nk = norm_key(c)
        if nk in m: return m[nk]
        toks = [t for t in re.split(r'[^a-z0-9]+',nk) if t]
        for k,v in m.items():
            if all(t in k for t in toks): return v
    return None

def status(api_v, bm_v, thr=0.80):
    if api_v is None: return "benchmark"
    if bm_v is None: return "neu"
    if bm_v == 0: return "api"
    try:
        r = float(api_v)/float(bm_v)
        return "api" if r>=thr else "gemischt" if r>=0.40 else "benchmark"
    except: return "benchmark"

def score_s4(lp): return round(min(float(lp)/4.0*10,10),2) if lp is not None else None
def score_s5(mbit): return round(min(float(mbit)/600*10,10),2) if mbit is not None else None
def score_s6(dbm): return round(max(0,min((float(dbm)+110)/50*10,10)),2) if dbm is not None else None
def score_s6_pct(pct): return round(float(pct)/100*10,2) if pct is not None else None
def score_s7(rg,ab2): return round(min((float(rg or 0)/4*5)+(float(ab2 or 0)/100*5),10),2)
def score_s8(kd): return round(min(float(kd)/6*10,10),2) if kd is not None else None
def score_s3(pct): return round(float(pct)/100*10,2) if pct is not None else None
def score_s2(tp): return round(min(float(tp)/0.8*10,10),2) if tp is not None else None


# ─── XLSX/CSV PARSING ─────────────────────────────────────────────────────────
def _find_header(rows):
    """Findet echte Headerzeile (überspringt BNetzA-Metadaten)."""
    lat_tok = ["breitengrad","latitude","lat","wgs84breite"]
    lng_tok = ["laengengrad","longitude","lng","wgs84laenge"]
    gem_tok = ["gemeinde","gemeindename","ort","ortsname","kreisfreiestadt","verwaltungsgemeinschaft"]
    for i,row in enumerate(rows[:100]):
        if row is None: continue
        t = norm_key(" ".join(str(x or "") for x in (row if hasattr(row,"__iter__") else [row])))
        if any(x in t for x in lat_tok) and any(x in t for x in lng_tok): return i
        if "betreiber" in t and ("nennleistung" in t or "ladepunkt" in t): return i
        # Für S5/S6: Gemeinde-Spalte
        if any(x in t for x in gem_tok) and ("anteil" in t or "versorgung" in t or "4g" in t or "gbit" in t): return i
    return 0

def read_xlsx(data: bytes) -> Tuple[List[Dict], int]:
    """Liest XLSX, gibt (Zeilen, Header-Zeilenindex) zurück."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl fehlt: pip install openpyxl")
    with tempfile.NamedTemporaryFile(suffix=".xlsx",delete=False) as f:
        f.write(data); path=f.name
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = max(wb.worksheets, key=lambda s: s.max_row or 0)
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows: return [], 0
        hi = _find_header(all_rows)
        seen: Dict[str,int]={}
        headers=[]
        for i,h in enumerate(all_rows[hi] or []):
            n=str(h or "").strip() or f"_col_{i}"
            if n in seen: seen[n]+=1; n=f"{n}_{seen[n]}"
            else: seen[n]=0
            headers.append(n)
        rows=[]
        for row in all_rows[hi+1:]:
            if row is None or all(x is None or str(x).strip()=="" for x in row): continue
            rows.append({headers[i]:(row[i] if i<len(row) else None) for i in range(len(headers))})
        log(f"    XLSX: {len(rows):,} Zeilen, Header war Zeile {hi+1}")
        return rows, hi
    finally:
        try: os.unlink(path)
        except: pass

def read_csv_bytes(data: bytes) -> List[Dict]:
    for enc in ("utf-8-sig","cp1252","latin-1"):
        try:
            text = data.decode(enc,errors="replace")
            lines = text.splitlines()
            hi = _find_header([l.split(";") for l in lines[:100]])
            text2 = "\n".join(lines[hi:])
            for delim in (";","\t",",","|"):
                reader = csv.DictReader(io.StringIO(text2), delimiter=delim)
                rows = list(reader)
                if rows and len(rows[0])>=3:
                    nk = norm_key(" ".join(rows[0].keys()))
                    if any(x in nk for x in ["breitengrad","latitude","betreiber","nennleistung","gemeinde"]):
                        log(f"    CSV: {len(rows):,} Zeilen ({enc}, '{delim}')")
                        return rows
        except: continue
    raise RuntimeError("CSV unlesbar")


# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────
class _LP(HTMLParser):
    def __init__(self): super().__init__(); self.links=[]
    def handle_starttag(self,t,a):
        if t=="a":
            h=dict(a).get("href","")
            if h: self.links.append(h)

def _fetch(url, **kw):
    headers = kw.pop("headers",{})
    headers.setdefault("User-Agent",UA)
    return requests.get(url, headers=headers, timeout=60, allow_redirects=True, **kw)

def _parse_bytes(data:bytes, url:str) -> List[Dict]:
    """Parst XLSX oder CSV Bytes."""
    # HTML-Zwischenseite → echten Link suchen
    if data[:200].lstrip().lower().startswith((b"<!doctype",b"<html")):
        p=_LP(); p.feed(data.decode("utf-8",errors="ignore"))
        for href in p.links:
            abs_url = urljoin(url,href)
            lu = abs_url.lower()
            if any(x in lu for x in [".xlsx",".csv","publicationfile"]):
                log(f"    → Redirect → {abs_url[:70]}")
                r2=_fetch(abs_url); r2.raise_for_status()
                return _parse_bytes(r2.content, abs_url)
        raise RuntimeError("HTML-Zwischenseite — kein Datei-Link gefunden")
    if data[:2]==b"PK":
        rows,_ = read_xlsx(data)
        return rows
    return read_csv_bytes(data)

def download_file(urls: List[str], page_url: str="", label: str="") -> Tuple[bytes, str]:
    """
    Lädt Datei von einer der URLs. Wenn alle fehlschlagen:
    → versucht Downloadseite zu parsen und aktuellen Link zu finden.
    Gibt (bytes, quelle_url) zurück.
    """
    errors = []
    # Direkte URLs
    for url in urls:
        try:
            log(f"    → {url[:75]}{'…' if len(url)>75 else ''}")
            r = _fetch(url)
            if r.ok and len(r.content) > 10000:
                log(f"    ✓ {len(r.content)//1024} KB geladen")
                return r.content, url
            errors.append(f"{url[:60]} → HTTP {r.status_code}")
        except Exception as e:
            errors.append(f"{url[:60]} → {str(e)[:50]}")
    # Fallback: Downloadseite durchsuchen
    if page_url:
        try:
            log(f"    → Suche Download-Link auf {page_url[:60]}")
            r = _fetch(page_url)
            if r.ok:
                p=_LP(); p.feed(r.text)
                candidates=[]
                for href in p.links:
                    abs_url = urljoin(page_url,href)
                    lu=abs_url.lower()
                    if any(x in lu for x in [".xlsx",".csv"]):
                        score = (3 if ".xlsx" in lu else 1)
                        # S5: bba_ im Namen
                        if "bba_" in lu or "breitband" in lu: score+=5
                        # S6: mobilfunk im Namen
                        if "mobilfunk" in lu and "auswertung" in lu: score+=5
                        candidates.append((score,abs_url))
                for _,url in sorted(candidates,reverse=True)[:5]:
                    try:
                        r2=_fetch(url)
                        if r2.ok and len(r2.content)>10000:
                            log(f"    ✓ {url[:60]} — {len(r2.content)//1024} KB")
                            return r2.content, url
                    except: pass
        except Exception as e:
            errors.append(f"Seite {page_url[:50]} → {e}")
    raise RuntimeError(
        f"{label}: Alle Downloads fehlgeschlagen.\n" +
        "\n".join(f"  - {e}" for e in errors) +
        f"\n\nManuell herunterladen und mit --file übergeben."
    )


# ─── NOMINATIM + POINT-IN-POLYGON ─────────────────────────────────────────────
def _load_geo(): 
    try: return json.loads(GEO_CACHE.read_text()) if GEO_CACHE.exists() else {}
    except: return {}

def _save_geo(c): GEO_CACHE.write_text(json.dumps(c,ensure_ascii=False,indent=2))

def _close(r): return r+[r[0]] if r and r[0]!=r[-1] else r

def _norm_geo(g):
    if not g: return None
    if g.get("type")=="Feature": g=g.get("geometry") or {}
    t=g.get("type")
    if t=="Polygon":
        coords=[_close([[float(p[0]),float(p[1])] for p in ring if len(p)>=2]) for ring in g.get("coordinates",[]) if len(ring)>=4]
        return {"type":"Polygon","coordinates":coords} if coords else None
    if t=="MultiPolygon":
        polys=[[_close([[float(p[0]),float(p[1])] for p in ring if len(p)>=2]) for ring in poly if len(ring)>=4] for poly in g.get("coordinates",[])]
        polys=[p for p in polys if p]
        return {"type":"MultiPolygon","coordinates":polys} if polys else None
    return None

def _bbox(g):
    pts=([p for ring in g["coordinates"] for p in ring] if g["type"]=="Polygon"
         else [p for poly in g["coordinates"] for ring in poly for p in ring])
    xs,ys=[p[0] for p in pts],[p[1] for p in pts]
    return [min(xs),min(ys),max(xs),max(ys)]

def get_boundary(k, cache):
    key=f"{k['name']}|{k.get('lk','')}"
    if key in cache: return cache[key]
    for q in [f"{k['name']}, {k.get('lk','')}, Sachsen-Anhalt, Deutschland",
              f"Gemeinde {k['name']}, Sachsen-Anhalt, Deutschland"]:
        try:
            r=requests.get("https://nominatim.openstreetmap.org/search",
                           params={"format":"jsonv2","polygon_geojson":"1","limit":"8","q":q,"countrycodes":"de"},
                           headers={"User-Agent":UA,"Accept-Language":"de"},timeout=25)
            if not r.ok: time.sleep(NOMINATIM_DELAY); continue
            items=r.json()
            best=next((x for x in items if "sachsen-anhalt" in str(x.get("display_name","")).lower()
                       and x.get("geojson") and x.get("class")=="boundary"),None)
            if not best: best=next((x for x in items if x.get("geojson")),None)
            if not best: time.sleep(NOMINATIM_DELAY); continue
            geo=_norm_geo(best.get("geojson"))
            if not geo: time.sleep(NOMINATIM_DELAY); continue
            result={"geometry":geo,"bbox":_bbox(geo),"source":"Nominatim OSM",
                    "display_name":best.get("display_name",""),"osm_id":best.get("osm_id")}
            cache[key]=result; _save_geo(cache)
            time.sleep(NOMINATIM_DELAY)
            return result
        except Exception as e:
            warn(f"Nominatim {k['name']}: {e}")
            time.sleep(NOMINATIM_DELAY)
    cache[key]=None; _save_geo(cache); return None

def _pip_ring(lng,lat,ring):
    inside=False; j=len(ring)-1
    for i in range(len(ring)):
        xi,yi=ring[i][0],ring[i][1]; xj,yj=ring[j][0],ring[j][1]
        if (yi>lat)!=(yj>lat):
            if lng<(xj-xi)*(lat-yi)/((yj-yi) or 1e-12)+xi: inside=not inside
        j=i
    return inside

def _pip_poly(lng,lat,poly):
    if not _pip_ring(lng,lat,poly[0]): return False
    return not any(_pip_ring(lng,lat,h) for h in poly[1:])

def pip(lng,lat,geo):
    if geo["type"]=="Polygon": return _pip_poly(lng,lat,geo["coordinates"])
    if geo["type"]=="MultiPolygon": return any(_pip_poly(lng,lat,p) for p in geo["coordinates"])
    return False

def in_bbox(lng,lat,b,pad=0.003): return (b[0]-pad)<=lng<=(b[2]+pad) and (b[1]-pad)<=lat<=(b[3]+pad)


# ─── SICHT 4: LADEINFRASTRUKTUR ───────────────────────────────────────────────
@dataclass
class LP:
    id:str; name:str; lat:float; lng:float
    anzahl:int; leistung_kw:Optional[float]; art:str
    adresse:str; ort:str

def _norm_lp(rows):
    LAT=["Breitengrad","Latitude","lat","Y","WGS84_Breite","Breitengrad (WGS84)"]
    LNG=["Längengrad","Laengengrad","Longitude","lng","X","WGS84_Laenge","Längengrad (WGS84)"]
    CNT=["Anzahl Ladepunkte","Anzahl_Ladepunkte","Ladepunkte"]
    PWR=["Nennleistung Ladeeinrichtung","Nennleistung_Ladeeinrichtung","Anschlussleistung [kW]"]
    ART=["Art Ladeeinrichtung","Art_Ladeeinrichtung","Art"]
    BET=["Betreiber","Betreibername","Firma"]
    ADR=["Straße","Strasse","Adresse"]
    ORT=["Ort","Gemeinde","Stadt"]
    pts,skip=[],0
    for i,row in enumerate(rows,1):
        lat=pf(pick(row,LAT)); lng=pf(pick(row,LNG))
        if lat is None or lng is None: skip+=1; continue
        if not((47<=lat<=56.5)and(5<=lng<=16)):
            if (47<=lng<=56.5)and(5<=lat<=16): lat,lng=lng,lat
            else: skip+=1; continue
        n=max(1,int(pf(pick(row,CNT)) or 1))
        leis=pf(pick(row,PWR))
        art_raw=str(pick(row,ART) or "").lower()
        art="DC" if (leis and leis>=50) or "dc" in art_raw or "schnell" in art_raw else "AC"
        pts.append(LP(f"bnetza-{i}",str(pick(row,BET) or "").strip(),round(lat,7),round(lng,7),
                      n,leis,art,str(pick(row,ADR) or "").strip(),str(pick(row,ORT) or "").strip()))
    log(f"    Normalisiert: {len(pts):,} Punkte, {skip:,} übersprungen")
    return pts

def sync_s4(kommunen, local_file=None):
    log("\n[Sicht 4] Ladeinfrastruktur · BNetzA XLSX/CSV")
    results={}; geo_cache=_load_geo()
    all_pts=None; src=""

    if local_file:
        try:
            data=Path(local_file).expanduser().read_bytes()
            rows=_parse_bytes(data,f"file://{local_file}")
            all_pts=_norm_lp(rows); src=Path(local_file).name
            ok(f"Lokale Datei: {len(all_pts):,} Punkte")
        except Exception as e: warn(f"Lokale Datei: {e}")

    if all_pts is None:
        try:
            data,src=download_file(S4_URLS, label="S4 Ladesäulenregister")
            rows=_parse_bytes(data,src); all_pts=_norm_lp(rows)
            ok(f"Download: {len(all_pts):,} Punkte")
        except Exception as e:
            err(f"S4 Download: {e}")

    for k in kommunen:
        name=k["name"]; bm=BENCHMARK.get(name,{}).get(4)
        if all_pts is not None:
            boundary=get_boundary(k,geo_cache)
            if boundary:
                bb=boundary["bbox"]
                sel=[p for p in all_pts if in_bbox(p.lng,p.lat,bb) and pip(p.lng,p.lat,boundary["geometry"])]
                total=sum(p.anzahl for p in sel); dc=sum(p.anzahl for p in sel if p.art=="DC")
                ac=total-dc; pro=round(total/(k["ew"]/1000),3) if k["ew"]>0 else 0
                st=status(pro,bm); sc=score_s4(pro)
                results[name]={"api":{"lp_pro_1000":pro,"total":total,"ac":ac,"dc":dc,
                               "stationen":len(sel),"boundary_ok":True,
                               "sample":[asdict(p) for p in sel[:50]]},
                               "benchmark":{"lp_pro_1000":bm,"quelle":"CSV 2024"},
                               "status":st,"score":sc,"quelle":f"BNetzA XLSX ({src[:40]})","via":"pip"}
                icon="✓" if st=="api" else "~" if st=="gemischt" else "⑊"
                log(f"    {icon} {name:30s} LP:{total:3d} ({pro:.2f}/1k) [{st}]")
                continue
        # Fallback Benchmark
        results[name]={"api":None,"benchmark":{"lp_pro_1000":bm,"quelle":"CSV 2024"},
                       "status":"benchmark","score":score_s4(bm) if bm else None,
                       "quelle":"Benchmark CSV 2024","via":"fallback"}
        fall(name)
    return results


# ─── SICHT 5: FESTNETZ / BREITBANDATLAS ───────────────────────────────────────
# XLSX-Struktur bba_12_2025.xlsx (verifiziert via gigabitgrundbuch.bund.de):
# Spalten u.a.: AGS, Gemeindenamen, Anteil_GBit_FTTB_H_%, Anteil_GBit_HFC_%,
#               Anteil_1000Mbit_%, Anteil_400Mbit_%, Anteil_100Mbit_%
# Wir nutzen: Anteil_100Mbit_% als Hauptkennzahl (vergleichbar mit Mbit/s-Benchmark)

S5_GEM_COLS  = ["Gemeindenamen","Gemeindename","Gemeinde","ort","GemName"]
S5_VAL_COLS  = ["Anteil_100Mbit_%","Anteil_100Mbit_Prozent","100 Mbit/s Anteil",
                "Anteil_GBit_FTTB_H_%","Anteil_GBit_HFC_%","Anteil_GBit_%"]
S5_MBIT_COLS = ["Mittelwert","Durchschnitt","avg_download","Download_Mbit"]

def sync_s5(kommunen, local_file=None):
    log("\n[Sicht 5] Festnetz · BNetzA Breitbandatlas XLSX")
    results={}; rows=None; src=""

    if local_file:
        try:
            data=Path(local_file).expanduser().read_bytes()
            rows,_=read_xlsx(data); src=Path(local_file).name
            ok(f"Lokal: {len(rows):,} Zeilen")
        except Exception as e: warn(f"Lokal: {e}")

    if rows is None:
        try:
            data,src=download_file(S5_URLS, S5_PAGE, "S5 Breitbandatlas")
            rows,_=read_xlsx(data)
            ok(f"Download: {len(rows):,} Zeilen")
        except Exception as e:
            err(f"S5: {e}"); rows=None

    # Lookup-Dict aufbauen: Gemeindename → Versorgungswert
    lookup: Dict[str,float] = {}
    if rows:
        for row in rows:
            gem = str(pick(row, S5_GEM_COLS) or "").strip()
            if not gem: continue
            # Primär: 100-Mbit/s-Versorgungsanteil
            val = pf(pick(row, S5_VAL_COLS))
            if val is None:
                # Fallback: Mbit/s-Durchschnitt
                val = pf(pick(row, S5_MBIT_COLS))
            if val is not None:
                lookup[gem.lower()] = val
                lookup[norm_key(gem)] = val
        ok(f"Breitband-Lookup: {len(lookup)//2} Gemeinden")

    for k in kommunen:
        name=k["name"]; bm=BENCHMARK.get(name,{}).get(5)
        api_val=None
        if lookup:
            # Exakter Match → normalisierter Match
            api_val = (lookup.get(name.lower()) or
                       lookup.get(norm_key(name)) or
                       next((v for key,v in lookup.items() if name.lower()[:8] in key),None))
        st=status(api_val,bm,thr=0.80 if api_val and api_val<=100 else 0.60)
        # Score: wenn Prozentwert → direkt; wenn Mbit/s → über 600er-Formel
        if api_val is not None:
            sc = score_s3(api_val) if api_val<=100 else score_s5(api_val)
        else:
            sc = score_s5(bm) if bm else None
        results[name]={
            "api":{"wert":api_val,"einheit":"% 100Mbit/s Versorgung" if api_val and api_val<=100 else "Mbit/s"} if api_val else None,
            "benchmark":{"mbit_s":bm,"quelle":"CSV 2024"},
            "status":st,"score":sc,
            "quelle":f"BNetzA Breitbandatlas ({src[:30]})" if api_val else "Benchmark CSV 2024",
        }
        icon="✓" if st=="api" else "~" if st=="gemischt" else "⑊"
        log(f"    {icon} {name:30s} API:{str(api_val or '—'):>8} BM:{str(bm or '—'):>8}")
    return results


# ─── SICHT 6: MOBILFUNK ───────────────────────────────────────────────────────
# XLSX-Struktur Auswertung_Mobilfunkmonitoring.xlsx (verifiziert):
# Spalten: AGS, Gemeindenamen, 4G_Telekom_%, 4G_Vodafone_%, 4G_O2_%,
#          5G_Telekom_%, 5G_Vodafone_%, 5G_O2_%, breitbandig_%
# Kennzahl: Anteil Fläche mit mind. 1 Anbieter 4G → vergleichbar mit Benchmark-dBm

S6_GEM_COLS = ["Gemeindenamen","Gemeindename","Gemeinde","GemName"]
S6_4G_COLS  = ["breitbandig_%","breitbandig_pct","Flaeche_4G_%","4G_gesamt",
                "4G_mind1Anbieter","Versorgung_4G","LTE_%"]
S6_AB_COLS  = ["4G_Telekom_%","LTE_Telekom_%"]  # Einzelanbieter Fallback

def sync_s6(kommunen, local_file=None):
    log("\n[Sicht 6] Mobilfunk · BNetzA Mobilfunk-Monitoring XLSX")
    results={}; rows=None; src=""

    if local_file:
        try:
            data=Path(local_file).expanduser().read_bytes()
            rows,_=read_xlsx(data); src=Path(local_file).name
            ok(f"Lokal: {len(rows):,} Zeilen")
        except Exception as e: warn(f"Lokal: {e}")

    if rows is None:
        try:
            data,src=download_file(S6_URLS, S6_PAGE, "S6 Mobilfunk-Monitoring")
            rows,_=read_xlsx(data)
            ok(f"Download: {len(rows):,} Zeilen")
        except Exception as e:
            err(f"S6: {e}"); rows=None

    # Lookup: Gemeindename → 4G-Flächenversorgung %
    lookup: Dict[str,float] = {}
    if rows:
        for row in rows:
            gem=str(pick(row,S6_GEM_COLS) or "").strip()
            if not gem: continue
            val=pf(pick(row,S6_4G_COLS))
            # Normierung: Werte können als Dezimal (0.97) oder Prozent (97) kommen
            if val is not None:
                if val<=1.0: val=round(val*100,2)
                lookup[gem.lower()]=val; lookup[norm_key(gem)]=val
        ok(f"Mobilfunk-Lookup: {len(lookup)//2} Gemeinden")

    for k in kommunen:
        name=k["name"]
        bm=BENCHMARK.get(name,{}).get(6,{})
        bm_vals=[v for v in [bm.get("o2"),bm.get("voda"),bm.get("tele")] if v is not None]
        bm_avg=round(sum(bm_vals)/len(bm_vals),2) if bm_vals else None

        api_pct=None
        if lookup:
            api_pct=(lookup.get(name.lower()) or
                     lookup.get(norm_key(name)) or
                     next((v for key,v in lookup.items() if name.lower()[:8] in key),None))

        # Status: Vergleich 4G-% aus Monitoring vs Benchmark-dBm
        # Gute 4G-Versorgung ≥ 95% entspricht etwa dBm-Wert > -90
        if api_pct is not None and bm_avg is not None:
            # Umrechnung Benchmark-dBm → Versorgungsqualität
            bm_qual = "gut" if bm_avg > -90 else "schlecht"
            api_qual = "gut" if api_pct >= 90 else "schlecht"
            st = "api" if api_qual==bm_qual else "gemischt"
        elif api_pct is not None:
            st="neu"
        else:
            st="benchmark"

        sc = score_s6_pct(api_pct) if api_pct is not None else (score_s6(bm_avg) if bm_avg else None)
        results[name]={
            "api":{"pct_4g":api_pct,"einheit":"% Flächenversorgung 4G"} if api_pct is not None else None,
            "benchmark":{"data":bm,"avg_dbm":bm_avg,"quelle":"CSV 2024 (dBm)"},
            "status":st,"score":sc,
            "quelle":f"BNetzA Mobilfunk-Monitoring ({src[:30]})" if api_pct else "Benchmark CSV 2024",
        }
        icon="✓" if st in("api","neu") else "~" if st=="gemischt" else "⑊"
        log(f"    {icon} {name:30s} 4G:{str(api_pct or '—'):>6}% BM-Ø:{str(bm_avg or '—'):>8} dBm")
    return results


# ─── SICHT 3: STREETVIEW ──────────────────────────────────────────────────────
def sync_s3(kommunen):
    log("\n[Sicht 3] StreetView · Google Metadata API")
    results={}
    if not GOOGLE_KEY:
        warn("GOOGLE_MAPS_API_KEY fehlt → Benchmark-Fallback")
        log("    Key erstellen: https://console.cloud.google.com")
        log("    → APIs & Services → Credentials → API Key → Street View Static API aktivieren")
        for k in kommunen:
            bm=BENCHMARK.get(k["name"],{}).get(3)
            results[k["name"]]={"api":None,"benchmark":{"pct":bm},"status":"benchmark",
                                 "score":score_s3(bm) if bm else None,"quelle":"Benchmark CSV 2024"}
        return results

    def grid(lat,lng,r=5,n=20):
        dlat=r/111.0; dlng=r/(111.0*math.cos(math.radians(lat)))
        s=max(1,int(n**0.5)); pts=[]
        for i in range(s):
            for j in range(s):
                pts.append((lat+dlat*(i/(s-1 or 1)*2-1),lng+dlng*(j/(s-1 or 1)*2-1)))
        return pts[:n]

    for k in kommunen:
        name,lat,lng=k["name"],k["lat"],k["lng"]
        bm=BENCHMARK.get(name,{}).get(3)
        pts=grid(lat,lng,r=4,n=16); hits=0
        for p_lat,p_lng in pts:
            try:
                r=requests.get("https://maps.googleapis.com/maps/api/streetview/metadata",
                               params={"location":f"{p_lat},{p_lng}","radius":500,"key":GOOGLE_KEY},
                               headers={"User-Agent":UA},timeout=10)
                if r.ok and r.json().get("status")=="OK": hits+=1
            except: pass
            time.sleep(0.08)
        pct=round(hits/len(pts)*100,1); st=status(pct,bm); sc=score_s3(pct)
        results[name]={"api":{"pct":pct,"getestet":len(pts),"treffer":hits},
                       "benchmark":{"pct":bm,"quelle":"CSV 2024"},
                       "status":st,"score":sc,"quelle":"Google StreetView Metadata API"}
        icon="✓" if st=="api" else "~" if st=="gemischt" else "⑊"
        log(f"    {icon} {name:30s} {pct:5.1f}% ({hits}/{len(pts)} Punkte)")
    return results


# ─── SICHT 2+9+10: OSM OVERPASS ───────────────────────────────────────────────
def sync_osm(kommunen):
    log("\n[Sicht 2+9+10] OSM Overpass API")
    results={}
    for k in kommunen:
        name,lat,lng,ew=k["name"],k["lat"],k["lng"],k["ew"]
        query=f"""
[out:json][timeout:40];
(
  node["amenity"="charging_station"](around:12000,{lat},{lng});
  way["amenity"="charging_station"](around:12000,{lat},{lng});
  node["tourism"~"museum|attraction|viewpoint|artwork"](around:8000,{lat},{lng});
  way["tourism"~"museum|attraction|viewpoint"](around:8000,{lat},{lng});
  node["amenity"~"community_centre|social_centre"](around:8000,{lat},{lng});
  node["office"~"political_party|association|ngo"](around:8000,{lat},{lng});
  node["amenity"~"school|kindergarten"](around:8000,{lat},{lng});
  way["amenity"~"school|kindergarten"](around:8000,{lat},{lng});
  way["landuse"~"industrial|commercial"](around:8000,{lat},{lng});
);
out tags;
"""
        try:
            r=requests.post("https://overpass-api.de/api/interpreter",
                            data={"data":query},headers={"User-Agent":UA},timeout=50)
            if not r.ok: raise RuntimeError(f"HTTP {r.status_code}")
            els=r.json().get("elements",[])
            lp=tour=ver=sch=kita=gew=0
            for el in els:
                t=el.get("tags",{})
                am,tour_t,off,lu=t.get("amenity",""),t.get("tourism",""),t.get("office",""),t.get("landuse","")
                if am=="charging_station": lp+=int(t.get("capacity",1))
                if tour_t in("museum","attraction","viewpoint","artwork"): tour+=1
                if am in("community_centre","social_centre") or off in("political_party","association","ngo"): ver+=1
                if am=="school": sch+=1
                if am=="kindergarten": kita+=1
                if lu in("industrial","commercial"): gew+=1
            tp=round(tour/(ew/1000),2) if ew>0 else 0
            bm2=BENCHMARK.get(name,{}).get(2)
            results[name]={
                "s2":{"api":{"touren_pro_1000":tp,"pois":tour},
                      "benchmark":{"tp":bm2},"status":status(tp,bm2),"score":score_s2(tp)},
                "s9":{"api":{"vereine":ver},"status":"neu" if ver>0 else "benchmark"},
                "s10":{"api":{"schulen":sch,"kitas":kita,"gewerbe_ways":gew,"lp_osm":lp},
                       "status":"neu" if sch>0 else "benchmark"},
                "quelle":"OSM Overpass API",
            }
            ok(f"{name}: LP={lp} Tour={tp:.2f}/Tsd Ver={ver} Sch={sch} Kita={kita}")
        except Exception as e:
            warn(f"OSM {name}: {e}")
            bm2=BENCHMARK.get(name,{}).get(2)
            results[name]={
                "s2":{"api":None,"benchmark":{"tp":bm2},"status":"benchmark"},
                "s9":{"api":None,"status":"benchmark"},"s10":{"api":None,"status":"benchmark"},
                "quelle":"Benchmark CSV 2024",
            }
        time.sleep(0.5)
    return results


# ─── SICHT 7: OZG / DIGITALE SERVICES ────────────────────────────────────────
# Hinweis: api.ozg-dashboard.de existiert nicht als öffentliche REST-API (Stand 2026).
# Das "Dashboard Digitale Verwaltung" (digitale-verwaltung.de) bietet keine offene API.
# Realistischer Weg: Serviceportal Sachsen-Anhalt (service.sachsen-anhalt.de) durchsuchen
# oder OZG-Umsetzungsstand als Excel vom IT-Planungsrat herunterladen.
# Für den Prototyp: Benchmark-CSV 2024 als primäre Quelle.

def sync_s7(kommunen, local_file=None):
    log("\n[Sicht 7] Digitale Services · OZG (Benchmark-CSV)")
    log("    Hinweis: Keine öffentliche OZG-REST-API verfügbar (Stand 2026)")
    log("    Quelle: Benchmark CSV-Erhebung 2024 + optionale Datei")
    results={}

    # Optional: manuelle CSV/XLSX mit OZG-Daten einlesen
    ozg_lookup: Dict[str, Dict] = {}
    if local_file:
        try:
            data=Path(local_file).expanduser().read_bytes()
            rows=_parse_bytes(data,f"file://{local_file}")
            for row in rows:
                gem=str(pick(row,["Gemeinde","Kommune","Name","Gemeindenamen"]) or "").strip()
                rg=pf(pick(row,["Reifegrad","reifegrad_mittelwert","OZG_Reifegrad","Reifegradmittelwert"]))
                ab2=pf(pick(row,["ab_Stufe2","Prozentanteil_ab_Stufe2","Anteil_Stufe2_%"]))
                if gem and (rg or ab2):
                    ozg_lookup[gem.lower()]={"rg":rg,"ab2":ab2}
                    ozg_lookup[norm_key(gem)]={"rg":rg,"ab2":ab2}
            ok(f"OZG-Datei: {len(ozg_lookup)//2} Gemeinden")
        except Exception as e:
            warn(f"OZG-Datei: {e}")

    for k in kommunen:
        name=k["name"]
        bm=BENCHMARK.get(name,{}).get(7,{})
        bm_rg=bm.get("rg"); bm_ab2=bm.get("ab2")
        # API-Lookup (manuelle Datei)
        api_val=ozg_lookup.get(name.lower()) or ozg_lookup.get(norm_key(name))
        if api_val:
            st=status(api_val.get("rg"), bm_rg)
            sc=score_s7(api_val.get("rg",0), api_val.get("ab2",0))
            results[name]={"api":api_val,"benchmark":{"rg":bm_rg,"ab2":bm_ab2,"quelle":"CSV 2024"},
                           "status":st,"score":sc,"quelle":f"OZG-Datei ({Path(local_file).name})"}
            ok(f"{name}: Reifegrad {api_val.get('rg')} | {api_val.get('ab2')}% ab Stufe 2")
        else:
            sc=score_s7(bm_rg or 0, bm_ab2 or 0) if bm_rg else None
            results[name]={"api":None,"benchmark":{"rg":bm_rg,"ab2":bm_ab2,"quelle":"CSV 2024"},
                           "status":"benchmark","score":sc,"quelle":"Benchmark CSV 2024"}
            fall(name)
    return results


# ─── SICHT 8: SOCIAL MEDIA ────────────────────────────────────────────────────
def sync_s8(kommunen):
    log("\n[Sicht 8] Social Media · Meta Graph API")
    results={}
    if not META_TOKEN:
        warn("META_ACCESS_TOKEN fehlt → Benchmark-Fallback")
        log("    Token erstellen: https://developers.facebook.com")
        log("    → My Apps → Create App → Graph API Explorer → Access Token")
        for k in kommunen:
            bm=BENCHMARK.get(k["name"],{}).get(8)
            results[k["name"]]={"api":None,"benchmark":{"kd":bm,"quelle":"CSV 2024"},
                                  "status":"benchmark","score":score_s8(bm) if bm else None,
                                  "quelle":"Benchmark CSV 2024 (kein META_TOKEN)"}
        return results

    cutoff=datetime.now(timezone.utc).timestamp()-90*24*3600
    for k in kommunen:
        name,ew=k["name"],k["ew"]
        bm=BENCHMARK.get(name,{}).get(8)
        all_pages=[]
        for term in [name,name.split("/")[0],name.split(" ")[0]][:2]:
            try:
                r=requests.get("https://graph.facebook.com/v19.0/search",
                               params={"q":term,"type":"page",
                                       "fields":"name,fan_count,category,posts.limit(1){created_time}",
                                       "limit":10,"access_token":META_TOKEN},
                               headers={"User-Agent":UA},timeout=15)
                if r.ok: all_pages.extend(r.json().get("data",[]))
            except Exception as e: warn(f"Meta {name}: {e}")
        aktive=0
        for page in all_pages:
            for post in page.get("posts",{}).get("data",[]):
                try:
                    ts=datetime.fromisoformat(post.get("created_time","").replace("Z","+00:00")).timestamp()
                    if ts>cutoff: aktive+=1; break
                except: pass
        kd=round(aktive/(ew/1000),2) if ew>0 else 0
        st=status(kd,bm); sc=score_s8(kd)
        results[name]={"api":{"kd":kd,"aktive_kanaele":aktive},
                       "benchmark":{"kd":bm,"quelle":"CSV 2024"},
                       "status":st,"score":sc,"quelle":"Meta Graph API"}
        icon="✓" if st=="api" else "~" if st=="gemischt" else "⑊"
        log(f"    {icon} {name:30s} {aktive} Kanäle → KD {kd:.2f}")
    return results


# ─── SICHT 1: IGEK PDF ────────────────────────────────────────────────────────
def sync_s1(pdf_path:str) -> Dict:
    log("\n[Sicht 1] IGEK/ISEK · Claude API PDF-Parsing")
    if not ANTHROPIC_KEY:
        warn("ANTHROPIC_API_KEY fehlt"); return {}
    p=Path(pdf_path).expanduser()
    if not p.exists(): err(f"PDF nicht gefunden: {p}"); return {}
    try:
        import anthropic
        client=anthropic.Anthropic(api_key=ANTHROPIC_KEY)
        pdf_b64=base64.b64encode(p.read_bytes()).decode()
        msg=client.messages.create(
            model="claude-sonnet-4-6",max_tokens=1000,
            messages=[{"role":"user","content":[
                {"type":"document","source":{"type":"base64","media_type":"application/pdf","data":pdf_b64}},
                {"type":"text","text":"""Analysiere dieses kommunale Entwicklungsdokument.
Die 10 Benchmarking-Sichten: 1=Strategie/IGEK, 2=Digitaler Zwilling, 3=360°-StreetView,
4=E-Ladeinfrastruktur, 5=Festnetz/Breitband, 6=Mobilfunk, 7=OZG/Digitale Services,
8=Social Media, 9=Vernetzung, 10=Daseinsvorsorge.
Antworte NUR mit JSON (kein Markdown):
{"abgedeckte_sichten":[<Nummern>],"anzahl":<int>,"fundstellen":{<nr>:<Stelle oder null>,...}}"""}
            ]}])
        text=msg.content[0].text
        m=re.search(r'\{.*\}',text,re.DOTALL)
        if m:
            res=json.loads(m.group())
            ok(f"{res.get('anzahl','?')}/10 IGEK-Sichten gefunden"); return res
    except Exception as e: err(f"Claude API: {e}")
    return {}


# ─── SCORE-BERECHNUNG & JSON-AUSGABE ──────────────────────────────────────────
def build_result(k, sr):
    name=k["name"]; scores={}
    for sicht in range(1,9):
        d=sr.get(sicht,{}).get(name)
        if d:
            if sicht in(2,9,10): d=d.get("s2" if sicht==2 else "s9" if sicht==9 else "s10",{})
            sc=d.get("score")
            if sc is not None: scores[sicht]=sc
        if sicht not in scores:
            bm=BENCHMARK.get(name,{}).get(sicht)
            if sicht==4 and isinstance(bm,(int,float)): scores[sicht]=score_s4(bm)
            elif sicht==5 and isinstance(bm,(int,float)): scores[sicht]=score_s5(bm)
            elif sicht==6 and isinstance(bm,dict):
                vals=[v for v in [bm.get("o2"),bm.get("voda"),bm.get("tele")] if v]
                if vals: scores[sicht]=score_s6(sum(vals)/len(vals))
            elif sicht==7 and isinstance(bm,dict):
                if bm.get("rg"): scores[sicht]=score_s7(bm["rg"],bm.get("ab2",0))
            elif sicht==8 and isinstance(bm,(int,float)): scores[sicht]=score_s8(bm)
            elif sicht==3 and isinstance(bm,(int,float)): scores[sicht]=score_s3(bm)
            elif sicht==2 and isinstance(bm,(int,float)): scores[sicht]=score_s2(bm)
    s1=scores.get(1,5.0)
    plat=[scores[s] for s in(2,3,4,5,6,7) if s in scores]
    plat_avg=round(sum(plat)/len(plat),2) if plat else 4.0
    s8=scores.get(8,3.0)
    ges=round(0.3*s1+0.5*plat_avg+0.2*s8,2)
    sichten_out={}
    for sicht in range(1,9):
        d=sr.get(sicht,{}).get(name)
        if d:
            if sicht in(2,9,10):
                sichten_out[str(sicht)]=d.get("s2" if sicht==2 else "s9" if sicht==9 else "s10",{})
            else:
                sichten_out[str(sicht)]=d
    return {"name":name,"landkreis":k.get("lk"),"einwohner":k.get("ew"),
            "ags":k.get("ags"),
            "scores":{**{f"s{i}":round(scores[i],2) for i in range(1,9) if i in scores},
                      "strategie":round(s1,2),"plattform":plat_avg,"netzwerk":round(s8,2),"gesamt":ges},
            "sichten":sichten_out,"osm":sr.get("osm",{}).get(name)}

def write_json(sr, kommunen, out_path):
    out=[build_result(k,sr) for k in kommunen]
    api_n=sum(1 for k in out if any(
        (k["sichten"].get(str(s)) or {}).get("status")=="api" for s in range(1,9)))
    data={"schema":"kommunalspiegel.alle_sichten.v6","version":VERSION,
          "generated_at":datetime.now(timezone.utc).isoformat(),
          "statistik":{"kommunen_gesamt":len(out),"kommunen_mit_api_daten":api_n},
          "datenquellen":{
              "s4":{"name":"BNetzA Ladesäulenregister","typ":"XLSX-Download","rhythmus":"täglich",
                    "url":"https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/start.html",
                    "key":"kein Key nötig"},
              "s5":{"name":"BNetzA Breitbandatlas","typ":"XLSX-Download","rhythmus":"halbjährlich",
                    "url":"https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/start.html",
                    "key":"kein Key nötig"},
              "s6":{"name":"BNetzA Mobilfunk-Monitoring","typ":"XLSX-Download","rhythmus":"halbjährlich",
                    "url":"https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/Auswertung_Mobilfunkmonitoring.xlsx",
                    "key":"kein Key nötig"},
              "s3":{"name":"Google StreetView Metadata","typ":"REST API","rhythmus":"monatlich",
                    "url":"https://maps.googleapis.com/maps/api/streetview/metadata",
                    "key":"GOOGLE_MAPS_API_KEY (kostenlos, console.cloud.google.com)"},
              "s2_9_10":{"name":"OSM Overpass API","typ":"REST API","rhythmus":"täglich",
                         "url":"https://overpass-api.de/api/interpreter","key":"kein Key nötig"},
              "s7":{"name":"OZG-Umsetzungsstand","typ":"Benchmark-CSV + opt. manuelle Datei",
                    "key":"kein Key nötig"},
              "s8":{"name":"Meta Graph API","typ":"REST API","rhythmus":"wöchentlich",
                    "url":"https://graph.facebook.com/v19.0","key":"META_ACCESS_TOKEN (developers.facebook.com)"},
              "s1":{"name":"Claude API PDF-Parsing","typ":"LLM API","rhythmus":"jährlich/manuell",
                    "url":"https://api.anthropic.com/v1/messages","key":"ANTHROPIC_API_KEY (console.anthropic.com)"},
          },
          "kommunen":out}
    Path(out_path).parent.mkdir(parents=True,exist_ok=True)
    Path(out_path).write_text(json.dumps(data,ensure_ascii=False,indent=2),encoding="utf-8")
    log(f"\n✓ JSON: {out_path}  ({Path(out_path).stat().st_size//1024} KB)")

def push_supabase(sr, kommunen):
    if not SUPABASE_URL: log("  Supabase nicht konfiguriert — nur lokale JSON"); return
    headers={"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}",
             "Content-Type":"application/json","Prefer":"resolution=merge-duplicates"}
    rows=[]
    for k in kommunen:
        name=k["name"]
        for sicht in range(1,9):
            d=sr.get(sicht,{}).get(name)
            if not d: continue
            rows.append({"kommune_name":name,"sicht_nr":sicht,"status":d.get("status","benchmark"),
                         "score":d.get("score"),"quelle":d.get("quelle",""),
                         "api_daten":json.dumps(d.get("api")),"benchmark_daten":json.dumps(d.get("benchmark")),
                         "synced_at":datetime.now(timezone.utc).isoformat()})
    try:
        r=requests.post(f"{SUPABASE_URL}/rest/v1/benchmark_sync",headers=headers,json=rows,timeout=30)
        ok(f"Supabase: {len(rows)} Zeilen {'✓' if r.ok else '✗ '+str(r.status_code)}")
    except Exception as e: warn(f"Supabase: {e}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    ap=argparse.ArgumentParser(description="KommunalSpiegel Sync-Agent v6")
    ap.add_argument("--alle",action="store_true",help="Alle Sichten")
    ap.add_argument("--sicht",type=int,help="Nur diese Sicht (1–10)")
    ap.add_argument("--kommune",type=str,help="Nur diese Kommune")
    ap.add_argument("--file",type=str,help="Lokale XLSX/CSV (für S4, S5, S6 oder S7)")
    ap.add_argument("--pdf",type=str,help="PDF für Sicht 1 (IGEK)")
    ap.add_argument("--out",type=str,default=str(OUT_DIR/"alle_sichten.json"))
    args=ap.parse_args()

    log("="*65)
    log(f"KommunalSpiegel Sync-Agent v{VERSION}")
    log(f"Start: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
    log(f"Keys: Google={'✓' if GOOGLE_KEY else '✗'}  "
        f"Meta={'✓' if META_TOKEN else '✗'}  "
        f"Anthropic={'✓' if ANTHROPIC_KEY else '✗'}  "
        f"Supabase={'✓' if SUPABASE_URL else '✗'}")
    log("="*65)

    ziel=[k for k in KOMMUNEN if not args.kommune or k["name"].lower()==args.kommune.lower()]
    if not ziel: err(f"'{args.kommune}' nicht gefunden"); sys.exit(1)
    log(f"Ziel: {len(ziel)} Kommune(n)")

    sr:Dict={}; t0=time.time()

    if args.alle or args.sicht==4:  sr[4]=sync_s4(ziel,args.file)
    if args.alle or args.sicht==5:  sr[5]=sync_s5(ziel,args.file)
    if args.alle or args.sicht==6:  sr[6]=sync_s6(ziel,args.file)
    if args.alle or args.sicht==3:  sr[3]=sync_s3(ziel)
    if args.alle or args.sicht in(2,9,10): sr["osm"]=sync_osm(ziel)
    if args.alle or args.sicht==7:  sr[7]=sync_s7(ziel,args.file)
    if args.alle or args.sicht==8:  sr[8]=sync_s8(ziel)
    if args.sicht==1 and args.pdf:  sr[1]={"_result":sync_s1(args.pdf)}

    write_json(sr,ziel,args.out)
    push_supabase(sr,ziel)

    log(f"\n{'='*65}")
    log(f"✓ Fertig in {time.time()-t0:.1f}s · {len(ziel)} Kommunen")
    log(f"  JSON: {args.out}")
    log(f"{'='*65}")

if __name__=="__main__":
    main()
