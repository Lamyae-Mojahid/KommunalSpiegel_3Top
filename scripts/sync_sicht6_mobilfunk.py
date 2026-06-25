"""
sync_sicht6_mobilfunk.py - KommunalSpiegel Sicht 6
Mobilfunkabdeckung, automatisiert über die offizielle BNetzA-Auswertungsdatei
"Mobilfunkstatistik (bis Gemeinde)" — bereits von der Bundesnetzagentur selbst
auf Gemeindeebene vorberechnete, exakte Prozentwerte je Technologie UND je
Netzbetreiber. Quartalsweise aktualisiert.

Warum diese Quelle statt WMS-Bildanalyse:
- Exakte, amtlich berechnete Prozentsätze (keine eigene Pixel-/Geometrie-
  Schätzung mehr nötig).
- Bereits nach Netzbetreiber (Telekom, Vodafone, Telefónica, 1&1) aufgeschlüsselt
  — genau die "je Anbieter"-Anforderung der ursprünglichen Spezifikation.
- Bereits auf Gemeindeebene aggregiert (Spalte 'AGS', 'Name', 'Verwaltungsebene').

Datei: https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/Auswertung_Mobilfunkmonitoring.xlsx
Blatt "Fläche": eine Zeile je Verwaltungseinheit (Bund/Land/Kreis/Gemeinde),
Spalten je Technologie (2G/4G/5G*/5G SA) für "Alle MNO" + je Netzbetreiber.

Begriffsklärung (5G* vs. 5G SA):
- "5G*" = jegliche 5G-Verfügbarkeit, inkl. 5G NSA (Non-Standalone, nutzt
  bestehende 4G-Infrastruktur als Unterbau — der heute weit verbreitete Typ).
- "5G SA" (Standalone) = reines 5G-Netz ohne 4G-Abhängigkeit, fortschrittlicher
  und seltener.
"""

import os
import sys
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import requests

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
log = logging.getLogger(__name__)

SUPABASE_URL = os.environ.get('SUPABASE_URL', '')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_ROLE_KEY', '') or os.environ.get('SUPABASE_SERVICE_KEY', '')

XLSX_URL = 'https://gigabitgrundbuch.bund.de/GIGA/DE/Downloads_Suche/aktuell/Auswertung_Mobilfunkmonitoring.xlsx'
USER_AGENT = 'KommunalSpiegel-Hochschule-Merseburg/1.0'

COMMUNES = [
    {'name': 'Leuna',          'kommune_id': 1, 'ags': '15088205'},
    {'name': 'Querfurt',       'kommune_id': 2, 'ags': '15088305'},
    {'name': 'Bad Dürrenberg', 'kommune_id': 3, 'ags': '15088020'},
]

COL_ALLE_MNO = {'2G': 6, '4G': 7, '5G': 8, '5G_SA': 9}
OPERATORS = {
    'Telekom':    {'2G': 13, '4G': 14, '5G': 15, '5G_SA': 16},
    'Vodafone':   {'2G': 17, '4G': 18, '5G': 19, '5G_SA': 20},
    'Telefonica': {'2G': 21, '4G': 22, '5G': 23, '5G_SA': 24},
    '1&1':        {'2G': 25, '4G': 26, '5G': 27, '5G_SA': 28},
}


def download_xlsx(path: str = 'mobilfunkstatistik.xlsx') -> Optional[str]:
    log.info(f"Lade offizielle BNetzA-Datei: {XLSX_URL}")
    try:
        r = requests.get(XLSX_URL, headers={'User-Agent': USER_AGENT}, timeout=120)
        if r.status_code != 200:
            log.error(f"Download fehlgeschlagen: HTTP {r.status_code}")
            return None
        with open(path, 'wb') as f:
            f.write(r.content)
        log.info(f"Gespeichert: {path} ({len(r.content):,} bytes)")
        return path
    except Exception as e:
        log.error(f"Download-Fehler: {e}")
        return None


def extract_kommune_row(ws, ags: str):
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row and str(row[0]) == ags:
            return row
    return None


def pct(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return round(float(val) * 100, 1)
    except (TypeError, ValueError):
        return None


def process_all() -> List[Dict[str, Any]]:
    import openpyxl

    path = download_xlsx()
    if not path:
        return []

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if 'Fläche' not in wb.sheetnames:
        log.error(f"Erwartetes Blatt 'Fläche' nicht gefunden. Vorhanden: {wb.sheetnames}")
        return []
    ws = wb['Fläche']

    results = []
    for k in COMMUNES:
        row = extract_kommune_row(ws, k['ags'])
        if not row:
            log.warning(f"{k['name']} (AGS {k['ags']}): keine Zeile in der Datei gefunden — überspringe")
            continue

        alle_mno = {tech: pct(row[idx]) for tech, idx in COL_ALLE_MNO.items()}
        je_anbieter = {
            anbieter: {tech: pct(row[idx]) for tech, idx in cols.items()}
            for anbieter, cols in OPERATORS.items()
        }

        log.info(f"{k['name']}: 2G={alle_mno['2G']}% · 4G={alle_mno['4G']}% · "
                  f"5G={alle_mno['5G']}% · 5G-SA={alle_mno['5G_SA']}%")
        for anbieter, vals in je_anbieter.items():
            log.info(f"  {anbieter}: 2G={vals['2G']}% · 4G={vals['4G']}% · 5G={vals['5G']}%")

        results.append({
            'kommune': k['name'], 'kommune_id': k['kommune_id'], 'ags': k['ags'],
            'alle_mno': alle_mno, 'je_anbieter': je_anbieter,
        })

    return results


def push_to_supabase(results: List[Dict[str, Any]]) -> None:
    if not SUPABASE_URL or not SUPABASE_KEY:
        log.warning('Supabase nicht konfiguriert (SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY fehlen) — kein Push.')
        return
    try:
        from supabase import create_client
    except ImportError:
        log.warning("Python-Paket 'supabase' nicht installiert — kein Push. (pip install supabase)")
        return

    base_url = SUPABASE_URL.split('/rest/')[0] if '/rest/' in SUPABASE_URL else SUPABASE_URL
    sb = create_client(base_url, SUPABASE_KEY)
    jahr = datetime.now().year

    for res in results:
        kommune_id = res['kommune_id']
        name = res['kommune']
        alle = res['alle_mno']

        kennzahlen = [
            ('Mobilfunk Fläche 2G (%, alle Anbieter)', alle.get('2G')),
            ('Mobilfunk Fläche 4G (%, alle Anbieter)', alle.get('4G')),
            ('Mobilfunk Fläche 5G (%, alle Anbieter)', alle.get('5G')),
            ('Mobilfunk Fläche 5G Standalone (%, alle Anbieter)', alle.get('5G_SA')),
        ]
        for kennzahl, wert in kennzahlen:
            if wert is None:
                continue
            try:
                sb.table('benchmark').upsert({
                    'kommune_id': kommune_id, 'sicht_nr': 6, 'kennzahl': kennzahl,
                    'wert_num': wert, 'score_normiert': None, 'erhebungsjahr': jahr,
                    'quelle': 'BNetzA Mobilfunkstatistik (bis Gemeinde), gigabitgrundbuch.bund.de',
                    'quelle_typ': 'automatisch',
                    'letzter_abruf': datetime.now(timezone.utc).isoformat(),
                }, on_conflict='kommune_id,sicht_nr,kennzahl,erhebungsjahr').execute()
            except Exception as e:
                log.error(f"  ✗ {name}: Supabase-Fehler (benchmark, {kennzahl}): {e}")

        try:
            sb.table('mobilfunk_anbieter').delete().eq('kommune_id', kommune_id).eq('erhebungsjahr', jahr).execute()
        except Exception as e:
            log.warning(f"  ⚠ {name}: Löschen alter mobilfunk_anbieter fehlgeschlagen: {e}")

        rows = [{
            'kommune_id': kommune_id, 'anbieter': anbieter, 'erhebungsjahr': jahr,
            'pct_2g': vals.get('2G'), 'pct_4g': vals.get('4G'),
            'pct_5g': vals.get('5G'), 'pct_5g_sa': vals.get('5G_SA'),
        } for anbieter, vals in res['je_anbieter'].items()]
        try:
            sb.table('mobilfunk_anbieter').insert(rows).execute()
            log.info(f"  ✓ {name}: benchmark + {len(rows)} Anbieter-Zeilen → Supabase")
        except Exception as e:
            log.error(f"  ✗ {name}: Supabase-Fehler (mobilfunk_anbieter): {e}")


def main() -> int:
    results = process_all()
    if not results:
        log.error('Keine Ergebnisse — Abbruch.')
        return 1

    log.info('══ Résumé final ══')
    for res in results:
        a = res['alle_mno']
        log.info(f"  {res['kommune']}: 2G={a['2G']}% · 4G={a['4G']}% · 5G={a['5G']}% · 5G-SA={a['5G_SA']}%")

    push_to_supabase(results)
    return 0


if __name__ == '__main__':
    sys.exit(main())
