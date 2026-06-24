#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
KommunalSpiegel · Sicht 4 Ladeinfrastruktur · Backend v6

Ziel: belastbare API-/Datenlogik für Sicht 4.
Ablauf:
1) BNetzA-Ladesäulenliste automatisch finden oder lokale XLSX/CSV lesen
2) Tabellenkopf robust erkennen
3) Koordinaten normalisieren und validieren
4) Gemeindegrenze laden: optional GeoJSON-Datei -> Overpass AGS -> Nominatim-Suche
5) Point-in-Polygon
6) Benchmark-Vergleich
7) Frontend-kompatible JSON schreiben

Wichtig: Ohne echte Grenze werden keine Live-Punkte gezählt. Dann bleibt Benchmark aktiv.
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sys
import tempfile
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import urljoin

import requests
from kommunen_seed_loader import load_seed, seed_kommunen, seed_benchmark_s4, seed_benchmark_s2, seed_benchmark_s10

try:
    from supabase import create_client
except ImportError:
    create_client = None

VERSION = "6.0.0"
SCHEMA = "kommunalspiegel.sicht4_ladeinfrastruktur.v6"

# ── Supabase: feste kommune_id-Zuordnung für die 3 Pilotkommunen ────
KOMMUNE_IDS = {
    "Leuna": 1,
    "Querfurt": 2,
    "Bad Dürrenberg": 3,
}

BNETZA_PAGE_URLS = [
    "https://www.bundesnetzagentur.de/Ladesaeulenkarte",
    "https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/DownloadundKontakt.html",
]
BNETZA_DIRECT_URLS = [
    # Stand 2026: offizielle Datenplattform der Bundesnetzagentur, wird zuerst probiert
    "https://data.bundesnetzagentur.de/Bundesnetzagentur/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/Ladesaeulenregister_BNetzA_2026-04-22.xlsx",
    "https://data.bundesnetzagentur.de/Bundesnetzagentur/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/Ladesaeulenregister_BNetzA_2026-04-22.csv",
    # ältere/alternative CMS-Pfade
    "https://www.bundesnetzagentur.de/DE/Fachthemen/ElektrizitaetundGas/E-Mobilitaet/_DL/Ladesaeuleninfrastruktur.xlsx?__blob=publicationFile&v=30",
    "https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/Sachgebiete/Energie/Unternehmen_Institutionen/E_Mobilitaet/Ladesaeulenregister.xlsx?__blob=publicationFile",
    "https://www.bundesnetzagentur.de/SharedDocs/Downloads/DE/Sachgebiete/Energie/Unternehmen_Institutionen/E_Mobilitaet/Ladesaeulenregister.csv?__blob=publicationFile",
]

USER_AGENT = "KommunalSpiegel/6.0 Hochschulprojekt Sicht4 Ladeinfrastruktur"
TIMEOUT = 90
NOMINATIM_DELAY = 1.1

DATA_DIR = Path("data")
CACHE_DIR = DATA_DIR / "cache"
RAW_DIR = DATA_DIR / "raw"
OUT_FILE = DATA_DIR / "sicht4_ladeinfrastruktur.json"
GEO_CACHE = CACHE_DIR / "gemeindegrenzen.json"

# Für die Projektkommunen. AGS bei den beiden kritischen Nachbarn explizit gesetzt.
KOMMUNEN: List[Dict[str, Any]] = [
    {"name":"Aken", "lk":"Anhalt-Bitterfeld", "ew":7200, "lat":51.8533, "lng":12.0446},
    {"name":"Köthen", "lk":"Anhalt-Bitterfeld", "ew":26500, "lat":51.7510, "lng":11.9729},
    {"name":"Muldestausee", "lk":"Anhalt-Bitterfeld", "ew":7800, "lat":51.6895, "lng":12.3165},
    {"name":"Osternienburger Land", "lk":"Anhalt-Bitterfeld", "ew":8700, "lat":51.7203, "lng":12.1472},
    {"name":"Raguhn-Jeßnitz", "lk":"Anhalt-Bitterfeld", "ew":8200, "lat":51.7131, "lng":12.2861},
    {"name":"Sandersdorf-Brehna", "lk":"Anhalt-Bitterfeld", "ew":14200, "lat":51.6124, "lng":12.2280},
    {"name":"Südliches Anhalt", "lk":"Anhalt-Bitterfeld", "ew":16100, "lat":51.6580, "lng":11.9055},
    {"name":"Zerbst/Anhalt", "lk":"Anhalt-Bitterfeld", "ew":20700, "lat":51.9679, "lng":12.0898},
    {"name":"Zörbig", "lk":"Anhalt-Bitterfeld", "ew":8600, "lat":51.6249, "lng":12.1249},
    {"name":"Hohenmölsen", "lk":"Burgenlandkreis", "ew":9200, "lat":51.1588, "lng":12.0944},
    {"name":"Lützen", "lk":"Burgenlandkreis", "ew":8100, "lat":51.2566, "lng":12.1407},
    {"name":"Teuchern", "lk":"Burgenlandkreis", "ew":7100, "lat":51.1508, "lng":12.0175},
    {"name":"Allstedt", "lk":"Mansfeld-Südharz", "ew":8300, "lat":51.3979, "lng":11.4119},
    {"name":"Arnstein", "lk":"Mansfeld-Südharz", "ew":7900, "lat":51.5588, "lng":11.3758},
    {"name":"Eisleben", "lk":"Mansfeld-Südharz", "ew":23800, "lat":51.5294, "lng":11.5492},
    {"name":"Gerbstedt", "lk":"Mansfeld-Südharz", "ew":7300, "lat":51.6252, "lng":11.6141},
    {"name":"Hettstedt", "lk":"Mansfeld-Südharz", "ew":13400, "lat":51.6488, "lng":11.5022},
    {"name":"Mansfeld", "lk":"Mansfeld-Südharz", "ew":7700, "lat":51.5968, "lng":11.4632},
    {"name":"Seegebiet Mansfelder Land", "lk":"Mansfeld-Südharz", "ew":8400, "lat":51.5699, "lng":11.7044},
    {"name":"Südharz", "lk":"Mansfeld-Südharz", "ew":8600, "lat":51.5117, "lng":10.9705},
    {"name":"Bad Dürrenberg", "lk":"Saalekreis", "ew":9800, "lat":51.2965, "lng":12.0645},
    {"name":"Bad Lauchstädt", "lk":"Saalekreis", "ew":9100, "lat":51.3875, "lng":11.8714},
    {"name":"Braunsbedra", "lk":"Saalekreis", "ew":10600, "lat":51.2872, "lng":11.8843},
    {"name":"Kabelsketal", "lk":"Saalekreis", "ew":11200, "lat":51.4522, "lng":12.0089},
    {"name":"Landsberg", "lk":"Saalekreis", "ew":9700, "lat":51.5219, "lng":11.9877, "ags":"15088195"},
    {"name":"Leuna", "lk":"Saalekreis", "ew":14500, "lat":51.3286, "lng":12.0032, "ags":"15088205"},
    {"name":"Mücheln (Geiseltal)", "lk":"Saalekreis", "ew":8600, "lat":51.2994, "lng":11.7996},
    {"name":"Petersberg", "lk":"Saalekreis", "ew":13100, "lat":51.5428, "lng":11.9875},
    {"name":"Querfurt", "lk":"Saalekreis", "ew":12800, "lat":51.3803, "lng":11.5897},
    {"name":"Salzatal", "lk":"Saalekreis", "ew":8800, "lat":51.4610, "lng":11.7783},
    {"name":"Schkopau", "lk":"Saalekreis", "ew":9300, "lat":51.3860, "lng":12.0049},
    {"name":"Teutschenthal", "lk":"Saalekreis", "ew":8900, "lat":51.4589, "lng":11.7875},
    {"name":"Wettin-Löbejün", "lk":"Saalekreis", "ew":11800, "lat":51.5490, "lng":11.8855},
]

BENCHMARK_PRO_TAUSEND: Dict[str, Optional[float]] = {
    "Köthen":0.57, "Muldestausee":0.60, "Sandersdorf-Brehna":1.13, "Südliches Anhalt":0.00,
    "Hohenmölsen":0.30, "Allstedt":0.53, "Eisleben":0.84, "Gerbstedt":0.00,
    "Seegebiet Mansfelder Land":0.94, "Bad Dürrenberg":2.95, "Bad Lauchstädt":0.89,
    "Braunsbedra":0.89, "Kabelsketal":0.11, "Landsberg":1.02, "Leuna":3.59,
    "Mücheln (Geiseltal)":0.21, "Querfurt":0.504, "Salzatal":0.88, "Wettin-Löbejün":0.21,
}

@dataclass
class Ladepunkt:
    name: str
    betreiber: str
    adresse: str
    plz: str
    ort: str
    lat: float
    lng: float
    anzahl: int
    art: str
    leistung_kw: Optional[float]
    anschlussleistung_kw: Optional[float]
    steckertypen: str
    inbetriebnahme: str
    quelle: str = "BNetzA Ladesäulenregister"


def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def norm_key(s: Any) -> str:
    s = str(s or "").strip().lower()
    repl = {"ä":"ae", "ö":"oe", "ü":"ue", "ß":"ss", "é":"e", "è":"e", "á":"a"}
    for a, b in repl.items():
        s = s.replace(a, b)
    return re.sub(r"[^a-z0-9]+", "", s)


def parse_float_de(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip().replace("\u00a0", " ").replace(" ", "")
    if not s or s.lower() in {"-", "–", "nan", "none", "k.a.", "ka"}:
        return None
    s = re.sub(r"[^0-9,\.\-]", "", s)
    if not s or s in {"-", ".", ","}:
        return None
    if "," in s and "." in s:
        # 1.234,56 oder 1,234.56
        s = s.replace(".", "").replace(",", ".") if s.rfind(",") > s.rfind(".") else s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def is_html(data: bytes) -> bool:
    h = data[:2048].lstrip().lower()
    return h.startswith(b"<!doctype") or h.startswith(b"<html") or b"<title" in h


def is_zip_xlsx(data: bytes) -> bool:
    return data[:2] == b"PK" and b"xl/" in data[:100000]


class LinkParser(HTMLParser):
    def __init__(self):
        super().__init__(); self.links: List[str] = []
    def handle_starttag(self, tag, attrs):
        if tag.lower() == "a":
            d = dict(attrs)
            if d.get("href"):
                self.links.append(d["href"])


def discover_bnetza_urls() -> List[str]:
    found: List[str] = []
    for page in BNETZA_PAGE_URLS:
        try:
            r = requests.get(page, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
            if not r.ok:
                continue
            p = LinkParser(); p.feed(r.text)
            for href in p.links:
                full = urljoin(page, href)
                nk = norm_key(full)
                if ("ladesaeulen" in nk or "ladesaulen" in nk or "ladesaeulenregister" in nk) and ("xlsx" in nk or "csv" in nk):
                    # Die große Liste bevorzugen, nicht Statistik-Excel
                    if "zahlen" not in nk and "karte" not in nk:
                        found.append(full)
        except Exception as e:
            log(f"WARN Downloadseiten-Erkennung fehlgeschlagen ({page}): {e}")
    # XLSX vor CSV, Direktlinks plus entdeckte Links, Duplikate entfernen
    ordered = BNETZA_DIRECT_URLS + found
    ordered.sort(key=lambda u: (0 if u.lower().endswith(".xlsx") or ".xlsx" in u.lower() else 1, u))
    seen=set(); out=[]
    for u in ordered:
        if u not in seen:
            seen.add(u); out.append(u)
    return out


def download_bytes(url: str) -> bytes:
    r = requests.get(url, headers={"User-Agent": USER_AGENT, "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/csv,*/*"}, timeout=TIMEOUT, allow_redirects=True)
    r.raise_for_status()
    return r.content


def find_header_row(rows: List[Iterable[Any]]) -> int:
    best_idx, best_score = 0, -999
    for idx, row in enumerate(rows[:200]):
        cells = [str(x or "") for x in row]
        txt = norm_key(" ".join(cells))
        non_empty = sum(1 for c in cells if c.strip())
        score = 0
        if any(x in txt for x in ["breitengrad", "latitude", "wgs84breite", "geolat"]): score += 45
        if any(x in txt for x in ["laengengrad", "langengrad", "longitude", "wgs84laenge", "geolon", "lng"]): score += 45
        if any(x in txt for x in ["betreiber", "ladeeinrichtung", "ladepunkt", "nennleistung", "anschlussleistung"]): score += 20
        if any(x in txt for x in ["strasse", "hausnummer", "postleitzahl", "plz", "ort"]): score += 15
        if non_empty >= 8: score += 10
        if idx > 30: score -= 3
        if score > best_score:
            best_idx, best_score = idx, score
    return best_idx


def rows_to_dicts(all_rows: List[Iterable[Any]], header_idx: int) -> Tuple[List[Dict[str, Any]], List[str]]:
    headers_raw = list(all_rows[header_idx] or [])
    seen: Dict[str, int] = {}; headers: List[str] = []
    for i, h in enumerate(headers_raw):
        name = str(h or "").strip() or f"_spalte_{i+1}"
        base = name
        if base in seen:
            seen[base] += 1
            name = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
        headers.append(name)
    out=[]
    for row in all_rows[header_idx+1:]:
        if row is None or all(x is None or str(x).strip()=="" for x in row):
            continue
        out.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
    return out, headers


def coord_valid_count(rows: List[Dict[str, Any]], limit: int = 5000) -> Tuple[int, Optional[str], Optional[str], str]:
    cols = list(rows[0].keys()) if rows else []
    best = (0, None, None, "normal")
    for lat_col in cols:
        for lon_col in cols:
            if lat_col == lon_col:
                continue
            cnt = 0
            for r in rows[:limit]:
                a = parse_float_de(r.get(lat_col)); b = parse_float_de(r.get(lon_col))
                if a is None or b is None:
                    continue
                # normal: lat/lon Deutschland
                if 47.0 <= a <= 55.5 and 5.0 <= b <= 16.0:
                    cnt += 1
            if cnt > best[0]:
                best = (cnt, lat_col, lon_col, "normal")
            cnt = 0
            for r in rows[:limit]:
                a = parse_float_de(r.get(lat_col)); b = parse_float_de(r.get(lon_col))
                if a is None or b is None:
                    continue
                # swapped: lon/lat
                if 5.0 <= a <= 16.0 and 47.0 <= b <= 55.5:
                    cnt += 1
            if cnt > best[0]:
                best = (cnt, lat_col, lon_col, "swapped")
    return best


def read_xlsx(data: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl fehlt. Bitte: pip install openpyxl")
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(data); p = tmp.name
    try:
        wb = load_workbook(p, read_only=True, data_only=True)
        candidates=[]
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hidx = find_header_row(rows)
            dicts, headers = rows_to_dicts(rows, hidx)
            coord_cnt, lat_col, lon_col, mode = coord_valid_count(dicts)
            nk = norm_key(" ".join(headers))
            header_score = (45 if "breitengrad" in nk else 0) + (45 if "laengengrad" in nk or "langengrad" in nk else 0) + (20 if "betreiber" in nk else 0)
            candidates.append((coord_cnt, header_score, len(dicts), ws.title, hidx, dicts, headers, lat_col, lon_col, mode))
        if not candidates:
            raise RuntimeError("XLSX enthält keine lesbare Tabelle")
        candidates.sort(key=lambda x: (x[0], x[1], x[2]), reverse=True)
        coord_cnt, header_score, nrows, title, hidx, dicts, headers, lat_col, lon_col, mode = candidates[0]
        log(f"  XLSX: Blatt '{title}', Kopfzeile {hidx+1}, Datenzeilen {len(dicts):,}, Koordinatenprobe {coord_cnt:,}")
        return dicts, {"format":"xlsx", "sheet":title, "header_row":hidx+1, "headers":headers, "coord_probe_count":coord_cnt, "coord_probe_lat":lat_col, "coord_probe_lon":lon_col, "coord_probe_mode":mode}
    finally:
        try: os.unlink(p)
        except OSError: pass


def read_csv_bytes(data: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    encodings = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]
    delimiters = [";", ",", "\t", "|"]
    best = None
    for enc in encodings:
        try:
            text = data.decode(enc)
        except UnicodeDecodeError:
            continue
        # drop metadata lines before real header
        lines = text.splitlines()
        hidx = find_header_row([[c for c in re.split(r"[;,\t|]", line)] for line in lines[:200]])
        tail = "\n".join(lines[hidx:])
        for delim in delimiters:
            try:
                reader = csv.DictReader(io.StringIO(tail), delimiter=delim)
                rows = [dict(r) for r in reader if any((v or "").strip() for v in r.values())]
                if not rows:
                    continue
                cnt, lat_col, lon_col, mode = coord_valid_count(rows)
                score = cnt + len(rows)//10000
                if best is None or score > best[0]:
                    best = (score, rows, enc, delim, hidx+1, reader.fieldnames or [], cnt, lat_col, lon_col, mode)
            except Exception:
                continue
    if best is None:
        raise RuntimeError("CSV konnte nicht gelesen werden")
    score, rows, enc, delim, hrow, headers, cnt, lat_col, lon_col, mode = best
    log(f"  CSV: Encoding {enc}, Delimiter {repr(delim)}, Kopfzeile {hrow}, Datenzeilen {len(rows):,}, Koordinatenprobe {cnt:,}")
    return rows, {"format":"csv", "encoding":enc, "delimiter":delim, "header_row":hrow, "headers":headers, "coord_probe_count":cnt, "coord_probe_lat":lat_col, "coord_probe_lon":lon_col, "coord_probe_mode":mode}


def read_table_bytes(data: bytes, source_name: str) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    if is_html(data):
        raise RuntimeError(f"Download lieferte HTML statt Daten: {source_name}")
    if is_zip_xlsx(data) or source_name.lower().split("?")[0].endswith(".xlsx"):
        try:
            return read_xlsx(data)
        except Exception as xerr:
            # manche Server liefern CSV trotz xlsx-URL oder umgekehrt
            try:
                return read_csv_bytes(data)
            except Exception:
                raise xerr
    return read_csv_bytes(data)


def download_bnetza(file_path: Optional[str], url: Optional[str]) -> Tuple[List[Dict[str, Any]], str, Dict[str, Any]]:
    if file_path:
        p = Path(file_path).expanduser()
        log(f"Lese lokale BNetzA-Datei: {p}")
        data = p.read_bytes()
        rows, meta = read_table_bytes(data, str(p))
        return rows, str(p), meta
    urls = [url] if url else discover_bnetza_urls()
    last_errors=[]
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    for u in urls:
        try:
            log(f"BNetzA Download-Versuch: {u}")
            data = download_bytes(u)
            if len(data) < 10_000:
                raise RuntimeError(f"Download zu klein ({len(data)} Bytes)")
            # Cache Rohdownload zur Nachvollziehbarkeit
            suffix = ".xlsx" if is_zip_xlsx(data) or ".xlsx" in u.lower() else ".csv"
            raw_path = RAW_DIR / f"bnetza_ladesaeulen_latest{suffix}"
            raw_path.write_bytes(data)
            rows, meta = read_table_bytes(data, u)
            meta["raw_file"] = str(raw_path)
            return rows, u, meta
        except Exception as e:
            last_errors.append(f"{u}: {e}")
            log(f"  WARN: {e}")
    raise RuntimeError("Kein BNetzA CSV/XLSX-Kandidat konnte gelesen werden. Letzte Fehler: " + " | ".join(last_errors[-5:]))


def choose_column(cols: List[str], aliases: List[str]) -> Optional[str]:
    nmap = {norm_key(c): c for c in cols}
    for a in aliases:
        na = norm_key(a)
        if na in nmap:
            return nmap[na]
    # fuzzy contains, aber keine zu kurzen Alias-Wörter wie "Ort" oder "Name"
    for c in cols:
        nc = norm_key(c)
        for a in aliases:
            na = norm_key(a)
            if len(na) >= 4 and (na in nc or nc in na):
                return c
    return None


def get_val(row: Dict[str, Any], cols: List[str], aliases: List[str], default: str = "") -> str:
    c = choose_column(cols, aliases)
    if c is None:
        return default
    v = row.get(c)
    return "" if v is None else str(v).strip()


def normalize_ladepunkte(rows: List[Dict[str, Any]], meta: Dict[str, Any], debug_columns: bool=False) -> Tuple[List[Ladepunkt], Dict[str, Any]]:
    if not rows:
        return [], {"error":"keine Zeilen"}
    cols = list(rows[0].keys())
    cnt, lat_col, lon_col, mode = coord_valid_count(rows, limit=min(10000, len(rows)))
    if not lat_col or not lon_col or cnt == 0:
        return [], {"error":"keine Koordinatenspalten erkannt", "columns":cols, **meta}
    aliases = {
        "betreiber": ["Betreiber", "Betreibername", "Anzeigename Betreiber", "Firma", "Anbieter"],
        "strasse": ["Straße", "Strasse", "Straße und Hausnummer", "Standort Straße", "Adresse"],
        "hausnr": ["Hausnummer", "Hausnr", "Nr."],
        "plz": ["PLZ", "Postleitzahl"],
        "ort": ["Ort", "Standortort", "Gemeinde", "Stadt"],
        "anzahl": ["Anzahl Ladepunkte", "Anzahl der Ladepunkte", "Ladepunkte", "Anzahl Ladepunkte je Ladeeinrichtung"],
        "leistung": ["Nennleistung Ladeeinrichtung", "Nennleistung", "Leistung", "max. Leistung", "Maximale Leistung [kW]", "Ladeleistung [kW]"],
        "anschluss": ["Anschlussleistung", "Netzanschlussleistung", "Anschlussleistung [kW]"],
        "stecker": ["Steckertypen", "Steckertyp", "Anschluss", "Anschlusstyp", "Stecker"],
        "inbetrieb": ["Inbetriebnahmedatum", "Inbetriebnahme", "Datum der Inbetriebnahme"],
        "name": ["Ladeeinrichtung", "Standortbezeichnung", "Name", "Bezeichnung", "Ladesäule"],
    }
    out: List[Ladepunkt] = []
    no_coord=0; outside=0; swapped_count=0
    for idx, r in enumerate(rows):
        a = parse_float_de(r.get(lat_col)); b = parse_float_de(r.get(lon_col))
        if a is None or b is None:
            no_coord += 1; continue
        lat, lng = (b, a) if mode == "swapped" else (a, b)
        if mode == "swapped": swapped_count += 1
        if not (47.0 <= lat <= 55.5 and 5.0 <= lng <= 16.0):
            outside += 1; continue
        betreiber = get_val(r, cols, aliases["betreiber"])
        strasse = get_val(r, cols, aliases["strasse"])
        hausnr = get_val(r, cols, aliases["hausnr"])
        plz = get_val(r, cols, aliases["plz"])
        ort = get_val(r, cols, aliases["ort"])
        name = get_val(r, cols, aliases["name"]) or betreiber or "Ladeeinrichtung"
        leistung = parse_float_de(get_val(r, cols, aliases["leistung"]))
        anschluss = parse_float_de(get_val(r, cols, aliases["anschluss"]))
        stecker = get_val(r, cols, aliases["stecker"])
        inbetrieb = get_val(r, cols, aliases["inbetrieb"])
        anz_raw = parse_float_de(get_val(r, cols, aliases["anzahl"]))
        anzahl = max(1, int(round(anz_raw))) if anz_raw is not None else 1
        art = "DC" if (leistung is not None and leistung >= 50) or re.search(r"\b(CCS|CHAdeMO|DC|HPC)\b", stecker or "", re.I) else "AC"
        adresse = " ".join(x for x in [strasse, hausnr] if x).strip()
        out.append(Ladepunkt(name=name, betreiber=betreiber, adresse=adresse, plz=plz, ort=ort, lat=round(lat, 7), lng=round(lng, 7), anzahl=anzahl, art=art, leistung_kw=leistung, anschlussleistung_kw=anschluss, steckertypen=stecker, inbetriebnahme=inbetrieb))
    diag = {
        **meta,
        "columns": cols,
        "lat_column": lat_col,
        "lon_column": lon_col,
        "coordinate_mode": mode,
        "rows_total": len(rows),
        "usable_coordinates": len(out),
        "rows_without_coordinates": no_coord,
        "rows_outside_germany": outside,
        "rows_swapped": swapped_count,
        "detected_columns": {
            "betreiber": choose_column(cols, aliases["betreiber"]),
            "name": choose_column(cols, aliases["name"]),
            "strasse": choose_column(cols, aliases["strasse"]),
            "hausnummer": choose_column(cols, aliases["hausnr"]),
            "plz": choose_column(cols, aliases["plz"]),
            "ort": choose_column(cols, aliases["ort"]),
            "anzahl_ladepunkte": choose_column(cols, aliases["anzahl"]),
            "leistung_kw": choose_column(cols, aliases["leistung"]),
            "anschlussleistung_kw": choose_column(cols, aliases["anschluss"]),
            "steckertypen": choose_column(cols, aliases["stecker"]),
            "inbetriebnahme": choose_column(cols, aliases["inbetrieb"]),
        }
    }
    if debug_columns:
        log("Spaltenerkennung:")
        for k, v in diag["detected_columns"].items():
            log(f"  {k:<24} -> {v}")
        log(f"  Koordinaten -> lat='{lat_col}', lon='{lon_col}', Modus={mode}, Probe={cnt}")
        log(f"  Verwertbare Koordinaten: {len(out):,} / {len(rows):,}")
    return out, diag

# GeoJSON helpers

def geom_bbox(geom: Dict[str, Any]) -> List[float]:
    xs=[]; ys=[]
    def walk(c):
        if isinstance(c, list) and c and isinstance(c[0], (int, float)):
            xs.append(float(c[0])); ys.append(float(c[1]))
        elif isinstance(c, list):
            for x in c: walk(x)
    walk(geom.get("coordinates", []))
    return [min(xs), min(ys), max(xs), max(ys)] if xs and ys else [0,0,0,0]


def normalize_geojson(obj: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not obj:
        return None
    if obj.get("type") == "FeatureCollection":
        feats = obj.get("features") or []
        if feats:
            return normalize_geojson(feats[0])
    if obj.get("type") == "Feature":
        return normalize_geojson(obj.get("geometry"))
    if obj.get("type") in {"Polygon", "MultiPolygon"}:
        return {"type": obj["type"], "coordinates": obj["coordinates"]}
    return None


def bbox_contains(bbox: List[float], lng: float, lat: float, pad: float = 0.0) -> bool:
    w,s,e,n = bbox
    return w-pad <= lng <= e+pad and s-pad <= lat <= n+pad


def load_geo_cache() -> Dict[str, Any]:
    if GEO_CACHE.exists():
        try: return json.loads(GEO_CACHE.read_text(encoding="utf-8"))
        except Exception: return {}
    return {}


def save_geo_cache(cache: Dict[str, Any]) -> None:
    GEO_CACHE.parent.mkdir(parents=True, exist_ok=True)
    GEO_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def fetch_boundary_overpass_ags(ags: str) -> Optional[Dict[str, Any]]:
    if not ags:
        return None
    q = f'''[out:json][timeout:60];
relation["boundary"="administrative"]["admin_level"="8"]["de:amtlicher_gemeindeschluessel"="{ags}"];
out tags;'''
    r = requests.post("https://overpass-api.de/api/interpreter", data={"data": q}, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
    r.raise_for_status()
    data = r.json()
    els = data.get("elements") or []
    rels = [e for e in els if e.get("type") == "relation" and e.get("id")]
    if not rels:
        return None
    rel_id = rels[0]["id"]
    time.sleep(NOMINATIM_DELAY)
    lookup = requests.get("https://nominatim.openstreetmap.org/lookup", params={"format":"jsonv2", "polygon_geojson":1, "osm_ids":f"R{rel_id}"}, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
    lookup.raise_for_status()
    items = lookup.json()
    if not items:
        return None
    geo = normalize_geojson(items[0].get("geojson"))
    if not geo:
        return None
    return {"geometry": geo, "bbox": geom_bbox(geo), "source": "OSM Relation via Overpass AGS + Nominatim lookup", "osm_relation_id": rel_id, "display_name": items[0].get("display_name"), "cached_at": datetime.now(timezone.utc).isoformat()}


def fetch_boundary_nominatim(k: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    queries = []
    if k.get("ags"):
        queries.append(k["ags"])
    queries += [f'{k["name"]}, {k.get("lk", "")}, Sachsen-Anhalt, Deutschland', f'Stadt {k["name"]}, Sachsen-Anhalt, Deutschland', f'{k["name"]}, Sachsen-Anhalt']
    for q in queries:
        r = requests.get("https://nominatim.openstreetmap.org/search", params={"format":"jsonv2", "polygon_geojson":1, "addressdetails":1, "limit":10, "q":q, "countrycodes":"de"}, headers={"User-Agent": USER_AGENT}, timeout=TIMEOUT)
        r.raise_for_status()
        items = r.json()
        best=None
        for item in items:
            addr = item.get("address") or {}
            if addr.get("country_code") != "de":
                continue
            geo = normalize_geojson(item.get("geojson"))
            if not geo:
                continue
            bb = geom_bbox(geo)
            if bbox_contains(bb, k["lng"], k["lat"], pad=0.02):
                best = item; break
        if best:
            geo = normalize_geojson(best.get("geojson"))
            return {"geometry":geo, "bbox":geom_bbox(geo), "source":"Nominatim OSM Polygon", "osm_id":best.get("osm_id"), "osm_type":best.get("osm_type"), "display_name":best.get("display_name"), "cached_at":datetime.now(timezone.utc).isoformat()}
        time.sleep(NOMINATIM_DELAY)
    return None


def fetch_boundary(k: Dict[str, Any], cache: Dict[str, Any], boundary_file: Optional[str] = None) -> Optional[Dict[str, Any]]:
    key = f'{k["name"]}|{k.get("ags", "")}'
    if boundary_file:
        p = Path(boundary_file).expanduser()
        obj = json.loads(p.read_text(encoding="utf-8"))
        geo = normalize_geojson(obj)
        if not geo:
            raise RuntimeError(f"Boundary-Datei enthält kein Polygon/MultiPolygon: {p}")
        res = {"geometry":geo, "bbox":geom_bbox(geo), "source":f"lokale GeoJSON-Datei: {p}", "display_name":k["name"], "cached_at":datetime.now(timezone.utc).isoformat()}
        return res
    if key in cache and cache[key]:
        return cache[key]
    log(f"Gemeindegrenze laden: {k['name']} …")
    try:
        res = fetch_boundary_overpass_ags(k.get("ags")) if k.get("ags") else None
        if res and bbox_contains(res["bbox"], k["lng"], k["lat"], pad=0.02):
            cache[key] = res; save_geo_cache(cache)
            log(f"  ✓ Grenze via AGS/OSM: {res.get('display_name')}")
            return res
    except Exception as e:
        log(f"  WARN Overpass/AGS: {e}")
    try:
        res = fetch_boundary_nominatim(k)
        if res and bbox_contains(res["bbox"], k["lng"], k["lat"], pad=0.02):
            cache[key] = res; save_geo_cache(cache)
            log(f"  ✓ Grenze via Nominatim: {res.get('display_name')}")
            return res
    except Exception as e:
        log(f"  WARN Nominatim: {e}")
    cache[key] = None; save_geo_cache(cache)
    log("  ✗ keine belastbare Gemeindegrenze")
    return None


def point_in_ring(lng: float, lat: float, ring: List[List[float]]) -> bool:
    inside = False
    j = len(ring) - 1
    for i in range(len(ring)):
        xi, yi = ring[i][0], ring[i][1]
        xj, yj = ring[j][0], ring[j][1]
        if (yi > lat) != (yj > lat):
            xint = (xj - xi) * (lat - yi) / ((yj - yi) or 1e-12) + xi
            if lng < xint:
                inside = not inside
        j = i
    return inside


def point_in_polygon(lng: float, lat: float, poly: List[Any]) -> bool:
    return bool(poly) and point_in_ring(lng, lat, poly[0]) and not any(point_in_ring(lng, lat, hole) for hole in poly[1:])


def point_in_geom(p: Ladepunkt, geo: Dict[str, Any]) -> bool:
    if geo["type"] == "Polygon":
        return point_in_polygon(p.lng, p.lat, geo["coordinates"])
    return any(point_in_polygon(p.lng, p.lat, poly) for poly in geo["coordinates"])


def in_bbox(p: Ladepunkt, bbox: List[float], pad: float = 0.002) -> bool:
    return bbox_contains(bbox, p.lng, p.lat, pad=pad)


def point_dict(p: Ladepunkt) -> Dict[str, Any]:
    d = asdict(p)
    d["anzahl_ladepunkte"] = p.anzahl
    d["leistung_kw"] = p.leistung_kw
    return d


def quality_status(total_lp: int, benchmark_pro_1000: Optional[float], ew: int, boundary_ok: bool) -> Tuple[str, Optional[float], str]:
    if not boundary_ok:
        return "benchmark", None, "Keine echte Gemeindegrenze verfügbar – Benchmark bleibt Basis"
    if total_lp == 0:
        return "benchmark", 0.0, "Keine BNetzA-Ladepunkte innerhalb der Gemeindegrenze – Benchmark bleibt Basis"
    if benchmark_pro_1000 is None:
        return "neu", None, "API-Daten vorhanden; kein Benchmark zum Vergleich vorhanden"
    expected_lp = benchmark_pro_1000 * ew / 1000.0
    if expected_lp <= 0:
        return "api", None, "Benchmark ist 0; API-Daten werden als aktueller Stand übernommen"
    ratio = total_lp / expected_lp
    if ratio >= 0.80:
        return "api", round(ratio, 3), f"API ersetzt Benchmark ({ratio:.0%} des Benchmark-Erwartungswerts)"
    if ratio >= 0.40:
        return "gemischt", round(ratio, 3), f"API teilweise ({ratio:.0%}); API + Benchmark gemeinsam anzeigen"
    return "benchmark", round(ratio, 3), f"API deutlich unter Benchmark ({ratio:.0%}); Benchmark bleibt Basis"


def aggregate(k: Dict[str, Any], all_points: List[Ladepunkt], boundary: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    bm = BENCHMARK_PRO_TAUSEND.get(k["name"])
    ew = int(k.get("ew") or 0)
    selected: List[Ladepunkt] = []
    if boundary:
        candidates = [p for p in all_points if in_bbox(p, boundary["bbox"])]
        selected = [p for p in candidates if point_in_geom(p, boundary["geometry"])]
    total = sum(p.anzahl for p in selected)
    dc = sum(p.anzahl for p in selected if p.art == "DC")
    ac = max(0, total - dc)
    pro1000 = round(total/(ew/1000), 3) if ew else 0
    status, ratio, msg = quality_status(total, bm, ew, boundary is not None)
    if status in {"api", "gemischt", "neu"} and ew:
        weighted = (ac * 0.3 + dc * 0.7) / (ew/1000)
        score = round(min(weighted / 4 * 10, 10), 2)
    elif bm is not None:
        score = round(min((bm * 0.65) / 4 * 10, 10), 2)
    else:
        score = None
    sample = [point_dict(p) for p in selected[:500]]
    return {
        "kommune": k["name"], "landkreis": k.get("lk"), "einwohner": ew, "ags": k.get("ags"),
        "_alle_ladepunkte": selected,
        "api": {"ladeeinrichtungen": len(selected), "ladepunkte_gesamt": total, "ladepunkte_ac": ac, "ladepunkte_dc": dc, "ladepunkte_pro_1000_ew": pro1000, "boundary_ok": boundary is not None, "boundary_source": boundary.get("source") if boundary else None, "osm_display_name": (boundary or {}).get("display_name")},
        "benchmark": {"ladepunkte_pro_1000_ew": bm, "quelle": "studentische manuelle Erhebung 2024/2025", "benchmark_ratio": ratio},
        "status": status, "status_message": msg, "score_s4": score,
        "points": sample, "ladepunkte_sample": sample,
        "boundary_ok": boundary is not None,
        "ladeeinrichtungen": len(selected),
        "ladepunkte_gesamt": total,
        "ladepunkte_ac": ac,
        "ladepunkte_dc": dc,
        "ladepunkte_pro_1000_ew": pro1000,
        "benchmark_ladepunkte_pro_1000_ew": bm,
    }


def make_fallback(error: str, diagnostics: Optional[Dict[str, Any]]=None) -> Dict[str, Any]:
    kommunen=[]
    for k in KOMMUNEN:
        bm = BENCHMARK_PRO_TAUSEND.get(k["name"])
        score = round(min(((bm or 0) * 0.65) / 4 * 10, 10), 2) if bm is not None else None
        kommunen.append({"kommune":k["name"], "landkreis":k.get("lk"), "einwohner":k.get("ew"), "ags":k.get("ags"), "api":{"ladeeinrichtungen":0,"ladepunkte_gesamt":0,"ladepunkte_ac":0,"ladepunkte_dc":0,"ladepunkte_pro_1000_ew":0,"boundary_ok":False,"boundary_source":None,"osm_display_name":None}, "benchmark":{"ladepunkte_pro_1000_ew":bm,"quelle":"studentische manuelle Erhebung 2024/2025","benchmark_ratio":None}, "status":"benchmark", "status_message":"API/Parsing nicht belastbar – Benchmark-Fallback aktiv", "score_s4":score, "points":[], "ladepunkte_sample":[], "boundary_ok":False, "ladepunkte_gesamt":0, "benchmark_ladepunkte_pro_1000_ew":bm})
    return {"schema":SCHEMA, "version":VERSION, "generated_at":datetime.now(timezone.utc).isoformat(), "source":{"name":"Benchmark-Fallback", "api_status":"fehler", "error":error, "download_pages":BNETZA_PAGE_URLS}, "diagnostics":diagnostics or {}, "statistik":{"kommunen_gesamt":len(kommunen), "kommunen_api":0, "kommunen_gemischt":0, "kommunen_benchmark":len(kommunen), "ladepunkte_api_gesamt":0, "verwertbare_ladepunkte":0}, "qualitaetsregeln":{"api":">=80% Benchmark", "gemischt":"40–80% Benchmark", "benchmark":"<40%, Fehler oder keine Grenze", "neu":"kein Benchmark vorhanden"}, "kommunen":kommunen}


def _json_safe(obj: Any) -> Any:
    """Entfernt interne, nicht JSON-serialisierbare Hilfsfelder (z. B. die
    vollständige Ladepunkt-Liste, die nur für den Supabase-Push gebraucht wird)."""
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items() if k != "_alle_ladepunkte"}
    if isinstance(obj, list):
        return [_json_safe(v) for v in obj]
    return obj


def write_json(path: Path, obj: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_json_safe(obj), ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"✓ JSON geschrieben: {path}")


def push_ladesaeulen_to_supabase(sb, kommune_id: int, name: str, ladepunkte: List[Ladepunkt], jahr: int) -> None:
    """Schreibt jede einzelne Ladesäule für eine Kommune in die Tabelle
    'ladesaeulen' — für die spätere Kartenansicht. Wie bei streetview_routes:
    alte Zeilen für (kommune_id, erhebungsjahr) zuerst löschen, dann neu
    einfügen, damit ein erneuter Lauf keine Duplikate erzeugt.
    """
    try:
        sb.table("ladesaeulen").delete().eq("kommune_id", kommune_id).eq("erhebungsjahr", jahr).execute()
    except Exception as e:
        log(f"  ⚠ {name}: Löschen alter ladesaeulen-Zeilen fehlgeschlagen: {e}")

    if not ladepunkte:
        return

    rows = [{
        "kommune_id": kommune_id,
        "betreiber": p.betreiber or None,
        "adresse": p.adresse or None,
        "plz": p.plz or None,
        "ort": p.ort or None,
        "lat": p.lat,
        "lng": p.lng,
        "anzahl_ladepunkte": p.anzahl,
        "art": p.art or None,
        "leistung_kw": p.leistung_kw,
        "steckertypen": p.steckertypen or None,
        "inbetriebnahme": p.inbetriebnahme or None,
        "erhebungsjahr": jahr,
    } for p in ladepunkte]

    batch_size = 500
    gespeichert = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i + batch_size]
        try:
            sb.table("ladesaeulen").insert(batch).execute()
            gespeichert += len(batch)
            log(f"  {name}: {gespeichert}/{len(rows)} Ladesäulen gespeichert")
        except Exception as e:
            log(f"  ✗ {name}: Fehler beim Speichern von Ladesäulen (Batch {i}): {e}")


def push_to_supabase(summaries: List[Dict[str, Any]]) -> None:
    """Schreibt die Sicht-4-Ergebnisse (Ladepunkte pro 1000 EW) als neue,
    automatisch erzeugte Zeile in die Supabase-Tabelle 'benchmark'.
    Bestehende manuelle Zeilen (quelle_typ='manuell', erhebungsjahr=2025)
    bleiben unverändert — es wird nur eine Zeile mit erhebungsjahr=aktuelles
    Jahr und quelle_typ='automatisch' hinzugefügt/aktualisiert.
    """
    url = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY")
    if not url or not key:
        log("Supabase nicht konfiguriert (SUPABASE_URL/SUPABASE_SERVICE_ROLE_KEY fehlen) — kein Push.")
        return
    if create_client is None:
        log("Python-Paket 'supabase' nicht installiert — kein Push. (pip install supabase)")
        return

    sb = create_client(url, key)
    jahr = datetime.now().year
    gepusht = 0
    for s in summaries:
        name = s.get("kommune")
        kommune_id = KOMMUNE_IDS.get(name)
        if kommune_id is None:
            continue  # nicht eine unserer 3 Pilotkommunen
        pro1000 = s.get("ladepunkte_pro_1000_ew")
        if pro1000 is None:
            log(f"  ⚠ {name}: kein Live-Wert (Status={s.get('status')}) — überspringe Push")
            continue
        try:
            sb.table("benchmark").upsert({
                "kommune_id": kommune_id,
                "sicht_nr": 4,
                "kennzahl": "Ladepunkte pro Tsd. Einwohner für E-Autos",
                "wert_num": pro1000,
                "score_normiert": s.get("score_s4"),
                "erhebungsjahr": jahr,
                "quelle": "Bundesnetzagentur Ladesäulenregister",
                "quelle_typ": "automatisch",
                "letzter_abruf": datetime.now(timezone.utc).isoformat(),
            }, on_conflict="kommune_id,sicht_nr,kennzahl,erhebungsjahr").execute()
            gepusht += 1
            log(f"  ✓ {name}: {pro1000} Ladepunkte/1000 EW → Supabase")

            alle_punkte = s.get("_alle_ladepunkte", [])
            push_ladesaeulen_to_supabase(sb, kommune_id, name, alle_punkte, jahr)
        except Exception as e:
            log(f"  ✗ {name}: Supabase-Fehler: {e}")
    log(f"Supabase: {gepusht}/{len(KOMMUNE_IDS)} Pilotkommunen aktualisiert")


def main() -> None:
    global KOMMUNEN, BENCHMARK_PRO_TAUSEND
    ap = argparse.ArgumentParser(description="KommunalSpiegel Sicht 4 BNetzA Backend v6")
    ap.add_argument("--seed", default="data/kommunen_seed.json", help="zentrale Kommunen-/Benchmark-Datei")
    ap.add_argument("--out", default=str(OUT_FILE))
    ap.add_argument("--kommune", help="optional nur eine Kommune, z. B. Leuna")
    ap.add_argument("--file", help="lokale BNetzA XLSX/CSV")
    ap.add_argument("--download-url", help="expliziter Download-Link")
    ap.add_argument("--boundary-file", help="lokale Gemeindegrenze als GeoJSON, besonders nützlich für Offline-Tests")
    ap.add_argument("--debug-columns", action="store_true")
    args = ap.parse_args()

    DATA_DIR.mkdir(exist_ok=True); CACHE_DIR.mkdir(parents=True, exist_ok=True)
    seed = load_seed(args.seed);
    if seed:
        loaded_kommunen = seed_kommunen(seed)
        loaded_benchmark = seed_benchmark_s4(seed)
        if loaded_kommunen: KOMMUNEN = loaded_kommunen
        if loaded_benchmark is not None: BENCHMARK_PRO_TAUSEND = loaded_benchmark
        log(f"Seed geladen: {len(KOMMUNEN)} Kommunen aus {args.seed}")
    out_path = Path(args.out)
    if args.kommune:
        targets = [k for k in KOMMUNEN if norm_key(k["name"]) == norm_key(args.kommune)]
        if not targets:
            log(f"FEHLER: Kommune nicht gefunden: {args.kommune}"); sys.exit(1)
    else:
        targets = KOMMUNEN

    log(f"KommunalSpiegel Sicht 4 · Backend v{VERSION}")
    log(f"Ziel: {len(targets)} Kommune(n)")
    log("-"*72)
    try:
        rows, source_url, table_meta = download_bnetza(args.file, args.download_url)
        log("-"*72)
        log("Normalisiere BNetzA-Datensätze …")
        points, diag = normalize_ladepunkte(rows, table_meta, debug_columns=args.debug_columns)
        diag["source_url"] = source_url
        if not points:
            write_json(out_path, make_fallback("Normalisierung ergab 0 verwertbare Ladepunkte", diag)); return
    except Exception as e:
        log(f"FEHLER: {e}")
        write_json(out_path, make_fallback(str(e))); return

    log("-"*72)
    cache = load_geo_cache()
    summaries=[]
    for k in targets:
        boundary = fetch_boundary(k, cache, args.boundary_file)
        res = aggregate(k, points, boundary)
        summaries.append(res)
        log(f"{k['name']:<28} {res['ladepunkte_gesamt']:>4} LP · Status {res['status']:<9} · {res['status_message']}")
    api = sum(1 for s in summaries if s["status"] == "api")
    mix = sum(1 for s in summaries if s["status"] == "gemischt")
    bm = sum(1 for s in summaries if s["status"] == "benchmark")
    out = {"schema":SCHEMA, "version":VERSION, "generated_at":datetime.now(timezone.utc).isoformat(), "source":{"name":"Bundesnetzagentur Ladesäulenregister", "api_url":source_url, "download_pages":BNETZA_PAGE_URLS, "api_status":"ok", "note":"BNetzA-Datei gelesen; Punkte werden nur bei echter Gemeindegrenze per Point-in-Polygon als Live-Wert gezählt."}, "diagnostics":diag, "statistik":{"kommunen_gesamt":len(summaries), "kommunen_api":api, "kommunen_gemischt":mix, "kommunen_benchmark":bm, "ladepunkte_api_gesamt":sum(s["ladepunkte_gesamt"] for s in summaries), "bnetza_datensaetze":len(rows), "verwertbare_ladepunkte":len(points)}, "qualitaetsregeln":{"api":"API-Zählung >=80% Benchmark → API ersetzt Benchmark", "gemischt":"40–80% → beide anzeigen", "benchmark":"<40%, Fehler oder keine Grenze → Benchmark bleibt Basis", "neu":"kein Benchmark vorhanden → API allein"}, "kommunen":summaries}
    write_json(out_path, out)
    log(f"Fertig: API={api}, gemischt={mix}, benchmark={bm}, nutzbare BNetzA-Punkte={len(points):,}")

    log("-"*72)
    log("[Supabase] Push der Sicht-4-Ergebnisse für die 3 Pilotkommunen …")
    push_to_supabase(summaries)

if __name__ == "__main__":
    main()
